"""
Eco SR Reader — Versão Streamlit
Lê DICOM SR Philips EPIQ, preenche formulário editável por seções e exporta CSV/Excel.

Dependências:
    pip install streamlit pydicom openpyxl
"""

import io, os, re, math
import streamlit as st
import streamlit.components.v1

try:
    import pydicom
except ImportError:
    st.error("Execute: pip install pydicom"); st.stop()

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("Execute: pip install openpyxl"); st.stop()


# ═══════════════════════════════════════════════════════════════════════
# DEFINIÇÃO DO FORMULÁRIO
# ═══════════════════════════════════════════════════════════════════════

FORMULARIO = {
    "DADOS ANTROPOMÉTRICOS": [
        {"name": "Peso",             "unit": "kg",    "ref_mas": "", "ref_fem": "", "calc": False},
        {"name": "Altura",           "unit": "cm",    "ref_mas": "", "ref_fem": "", "calc": False},
        {"name": "Superfície corp.", "unit": "m²",    "ref_mas": "", "ref_fem": "", "calc": True},
        {"name": "IMC",              "unit": "kg/m²", "ref_mas": "", "ref_fem": "", "calc": True},
    ],
    "CÂMARAS ESQUERDAS": [
        {"name": "Aorta ascend.",         "unit": "mm",    "ref_mas": "",         "ref_fem": "",         "calc": False},
        {"name": "Aorta ascend. index",   "unit": "mm/m²", "ref_mas": "",         "ref_fem": "",         "calc": True},
        {"name": "Diâm VSVE",             "unit": "mm",    "ref_mas": "",         "ref_fem": "",         "calc": False},
        {"name": "Seio aórtico",          "unit": "mm",    "ref_mas": "31 - 37",  "ref_fem": "27 - 33",  "calc": False},
        {"name": "AE - Diâm.",            "unit": "mm",    "ref_mas": "30 - 40",  "ref_fem": "27 - 38",  "calc": False},
        {"name": "AE - Vol. bipl.",       "unit": "mL",    "ref_mas": "",         "ref_fem": "",         "calc": False},
        {"name": "AE - Vol. bipl. index", "unit": "mL/m²", "ref_mas": "16 - 34",  "ref_fem": "16 - 34",  "calc": True},
        {"name": "Septo",                 "unit": "mm",    "ref_mas": "6 - 10",   "ref_fem": "6 - 9",    "calc": False},
        {"name": "Parede post.",          "unit": "mm",    "ref_mas": "6 - 10",   "ref_fem": "6 - 9",    "calc": False},
        {"name": "DdVE",                  "unit": "mm",    "ref_mas": "42 - 58",  "ref_fem": "38 - 52",  "calc": False},
        {"name": "DsVE",                  "unit": "mm",    "ref_mas": "25 - 40",  "ref_fem": "22 - 35",  "calc": False},
        {"name": "VDVE",                  "unit": "mL",    "ref_mas": "62 - 150", "ref_fem": "46 - 106", "calc": True},
        {"name": "VSVE",                  "unit": "mL",    "ref_mas": "21 - 61",  "ref_fem": "14 - 42",  "calc": True},
        {"name": "VDVE index",            "unit": "mL/m²", "ref_mas": "24 - 74",  "ref_fem": "29 - 61",  "calc": True},
        {"name": "FEVE (Teichholz)",      "unit": "%",     "ref_mas": "52 - 72",  "ref_fem": "54 - 74",  "calc": True},
        {"name": "FEVE (Simpson)",        "unit": "%",     "ref_mas": "52 - 72",  "ref_fem": "54 - 74",  "calc": False},
        {"name": "ERP",                   "unit": "",      "ref_mas": "< 0,42",   "ref_fem": "< 0,42",   "calc": True},
        {"name": "Massa VE",              "unit": "g",     "ref_mas": "88 - 224", "ref_fem": "67 - 162", "calc": True},
        {"name": "Massa index",           "unit": "g/m²",  "ref_mas": "49 - 115", "ref_fem": "43 - 95",  "calc": True},
    ],
    "CÂMARAS DIREITAS": [
        {"name": "AD - Área",        "unit": "cm²",   "ref_mas": "",        "ref_fem": "",        "calc": False},
        {"name": "AD - Vol.",        "unit": "mL",    "ref_mas": "",        "ref_fem": "",        "calc": False},
        {"name": "AD - Vol. index",  "unit": "mL/m²", "ref_mas": "18 - 32", "ref_fem": "15 - 27", "calc": True},
        {"name": "AD - PSAP",        "unit": "mmHg",  "ref_mas": "<= 35",   "ref_fem": "<= 35",   "calc": False},
        {"name": "VD - Diâm. basal", "unit": "mm",    "ref_mas": "25 - 41", "ref_fem": "25 - 41", "calc": False},
        {"name": "VD - TAPSE",       "unit": "mm",    "ref_mas": ">= 17",   "ref_fem": ">= 17",   "calc": False},
        {"name": "VD - FAC",         "unit": "%",     "ref_mas": ">= 35",   "ref_fem": ">= 35",   "calc": False},
        {"name": "VD - Onda S",      "unit": "cm/s",  "ref_mas": ">= 9,5",  "ref_fem": ">= 9,5",  "calc": False},
        {"name": "VD - PSAP",        "unit": "mmHg",  "ref_mas": "<= 35",   "ref_fem": "<= 35",   "calc": False},
        {"name": "AP - Diâm.",       "unit": "mm",    "ref_mas": "",        "ref_fem": "",        "calc": False},
    ],
    "VALVA MITRAL": [
        {"name": "Vel. onda E",    "unit": "m/s",  "ref_mas": "",         "ref_fem": "",         "calc": False},
        {"name": "Vel. onda A",    "unit": "m/s",  "ref_mas": "",         "ref_fem": "",         "calc": False},
        {"name": "Relação E/A",    "unit": "",     "ref_mas": "",         "ref_fem": "",         "calc": True},
        {"name": "Decel. Time",    "unit": "ms",   "ref_mas": "150 - 250","ref_fem": "150 - 250","calc": False},
        {"name": "PHT",            "unit": "ms",   "ref_mas": "",         "ref_fem": "",         "calc": False},
        {"name": "Área (PHT)",     "unit": "cm²",  "ref_mas": "",         "ref_fem": "",         "calc": True},
        {"name": "Área (PISA)",    "unit": "cm²",  "ref_mas": "",         "ref_fem": "",         "calc": True},
        {"name": "VTI mit.",       "unit": "cm",   "ref_mas": "",         "ref_fem": "",         "calc": False},
        {"name": "Grad. máx.",     "unit": "mmHg", "ref_mas": "",         "ref_fem": "",         "calc": False},
        {"name": "Grad. méd.",     "unit": "mmHg", "ref_mas": "",         "ref_fem": "",         "calc": False},
        {"name": "MAPSE",          "unit": "mm",   "ref_mas": "",         "ref_fem": "",         "calc": False},
        {"name": "Dur. onda A",    "unit": "ms",   "ref_mas": "",         "ref_fem": "",         "calc": False},
        {"name": "Fluxo (PISA)",   "unit": "mL/s", "ref_mas": "",         "ref_fem": "",         "calc": True},
        {"name": "Volume Regurg.", "unit": "mL",   "ref_mas": "< 30",     "ref_fem": "< 30",     "calc": True},
    ],
    "AORTA / VSVE": [
        {"name": "Diâm VSVE",      "unit": "mm",    "ref_mas": "",        "ref_fem": "",        "calc": False},
        {"name": "VTI VSVE",       "unit": "cm",    "ref_mas": "",        "ref_fem": "",        "calc": False},
        {"name": "VTI Ao",         "unit": "cm",    "ref_mas": "",        "ref_fem": "",        "calc": False},
        {"name": "AVAo (EC-VTI)",  "unit": "cm²",   "ref_mas": "",        "ref_fem": "",        "calc": True},
        {"name": "AVAo (EC-Vmax)", "unit": "cm²",   "ref_mas": "",        "ref_fem": "",        "calc": True},
        {"name": "AVAo index",     "unit": "cm²/m²","ref_mas": "",        "ref_fem": "",        "calc": True},
        {"name": "Vel. Ratio",     "unit": "",      "ref_mas": "> 0,25",  "ref_fem": "> 0,25",  "calc": True},
        {"name": "Vel. máx. Ao",   "unit": "m/s",   "ref_mas": "< 2,0",   "ref_fem": "< 2,0",   "calc": False},
        {"name": "Grad. máx.",     "unit": "mmHg",  "ref_mas": "< 20",    "ref_fem": "< 20",    "calc": False},
        {"name": "Grad. méd.",     "unit": "mmHg",  "ref_mas": "< 10",    "ref_fem": "< 10",    "calc": False},
        {"name": "Decel. Slope",   "unit": "mm/s²", "ref_mas": "",        "ref_fem": "",        "calc": False},
        {"name": "PHT Ao",         "unit": "ms",    "ref_mas": "",        "ref_fem": "",        "calc": False},
    ],
    "TRICÚSPIDE / PULMONAR": [
        {"name": "Vel. RT",    "unit": "m/s",  "ref_mas": "",      "ref_fem": "",      "calc": False},
        {"name": "PSAP",       "unit": "mmHg", "ref_mas": "<= 35", "ref_fem": "<= 35", "calc": False},
        {"name": "AP - Diâm.", "unit": "mm",   "ref_mas": "",      "ref_fem": "",      "calc": False},
        {"name": "VTI AP",     "unit": "cm",   "ref_mas": "",      "ref_fem": "",      "calc": False},
    ],
    "TDI": [
        {"name": "Vel. e' septal",    "unit": "cm/s", "ref_mas": ">= 7",  "ref_fem": ">= 7",  "calc": False},
        {"name": "Rel. E/E' septal",  "unit": "",     "ref_mas": "<= 14", "ref_fem": "<= 14", "calc": True},
        {"name": "Vel. e' lateral",   "unit": "cm/s", "ref_mas": ">= 10", "ref_fem": ">= 10", "calc": False},
        {"name": "Rel. E/E' lateral", "unit": "",     "ref_mas": "<= 14", "ref_fem": "<= 14", "calc": True},
        {"name": "E/e' MÉDIO",        "unit": "",     "ref_mas": "<= 14", "ref_fem": "<= 14", "calc": True},
        {"name": "Vel. a' septal",    "unit": "cm/s", "ref_mas": "",      "ref_fem": "",      "calc": False},
        {"name": "Vel. a' lateral",   "unit": "cm/s", "ref_mas": "",      "ref_fem": "",      "calc": False},
        {"name": "E/A tecidual",      "unit": "",     "ref_mas": "",      "ref_fem": "",      "calc": True},
        {"name": "VP sistólica",       "unit": "m/s",  "ref_mas": "",      "ref_fem": "",      "calc": False},
        {"name": "VP diastólica",      "unit": "m/s",  "ref_mas": "",      "ref_fem": "",      "calc": False},
        {"name": "VP S/D",             "unit": "",     "ref_mas": "",      "ref_fem": "",      "calc": False},
        {"name": "TRIV",               "unit": "ms",   "ref_mas": "",      "ref_fem": "",      "calc": False},
    ],
    "STRAIN": [
        {"name": "VE - SLG",  "unit": "%", "ref_mas": "", "ref_fem": "", "calc": False},
        {"name": "VD - SLPL", "unit": "%", "ref_mas": "", "ref_fem": "", "calc": False},
        {"name": "AE - R",    "unit": "%", "ref_mas": "", "ref_fem": "", "calc": False},
        {"name": "AE - CD",   "unit": "%", "ref_mas": "", "ref_fem": "", "calc": False},
        {"name": "AE - B",    "unit": "%", "ref_mas": "", "ref_fem": "", "calc": False},
    ],
}


# ═══════════════════════════════════════════════════════════════════════
# ESTRUTURA DROPDOWNS
# ═══════════════════════════════════════════════════════════════════════

ESTRUTURA_DROPDOWNS = {
    "VENTRÍCULO ESQUERDO": {
        "Tamanho da cavidade": ["Normal", "Dilatação leve", "Dilatação moderada", "Dilatação importante"],
        "Geometria ventricular": ["Normal", "Remodelamento concêntrico", "Hipertrofia concêntrica", "Hipertrofia excêntrica"],
        "Função sistólica": ["Normal", "Reduzida de grau leve", "Reduzida de grau moderado", "Reduzida de grau importante"],
        "Função diastólica": ["Normal", "Disfunção diastólica grau I", "Disfunção diastólica grau II", "Disfunção diastólica grau III", "Função diastólica não avaliada"],
    },
    "VENTRÍCULO DIREITO": {
        "Tamanho da cavidade": ["Normal", "Dilatação leve", "Dilatação moderada", "Dilatação importante"],
        "Função sistólica": ["Normal", "Reduzida"],
    },
    "ÁTRIO ESQUERDO": {
        "Tamanho da cavidade": ["Normal", "Dilatação leve", "Dilatação moderada", "Dilatação importante"],
    },
    "ÁTRIO DIREITO": {
        "Tamanho da cavidade": ["Normal", "Dilatação leve", "Dilatação moderada", "Dilatação importante"],
    },
    "VALVA AORTA": {
        "Geral": ["Normal", "Calcificação", "Monocúspide", "Bicúspide", "Vegetação", "Prótese biológica", "Prótese mecânica"],
        "Estenose": ["Ausente", "Leve", "Moderada", "Importante"],
        "Refluxo": ["Ausente", "Leve", "Moderado", "Importante"],
    },
    "VALVA MITRAL": {
        "Geral": ["Normal", "Calcificação", "Reumática", "Mixomatosa", "Ruptura de cordoalha", "SAM", "Vegetação", "Prótese biológica", "Prótese mecânica"],
        "Estenose": ["Ausente", "Leve", "Moderada", "Importante"],
        "Refluxo": ["Ausente", "Leve", "Moderado", "Importante"],
    },
    "VALVA TRICÚSPIDE": {
        "Geral": ["Normal", "Calcificação", "Carcinóide", "Vegetação", "Prótese biológica", "Prótese mecânica"],
        "Estenose": ["Ausente", "Leve", "Moderada", "Importante"],
        "Refluxo": ["Ausente", "Leve", "Moderado", "Importante"],
    },
    "VALVA PULMONAR": {
        "Geral": ["Normal", "Calcificação", "Vegetação"],
        "Estenose": ["Ausente", "Leve", "Moderada", "Importante"],
        "Refluxo": ["Ausente", "Leve", "Moderado", "Importante"],
    },
    "AORTA": {
        "Raiz da aorta": ["Normal", "Dilatação"],
        "Aorta ascendente": ["Normal", "Dilatação"],
    },
    "ARTÉRIA PULMONAR": {
        "Tronco da pulmonar": ["Normal", "Dilatação"],
    },
    "PERICÁRDIO": {
        "Geral": ["Normal", "Espessado", "Derrame pericárdico leve", "Derrame pericárdico moderado", "Derrame pericárdico importante"],
    },
    "CONGÊNITAS": {
        "Geral": ["Ausente", "Comunicação interatrial (CIA)", "Comunicação interventricular (CIV)", "Persistência do canal arterial", "Forame Oval Patente (FOP)", "Transposição dos grandes vasos", "Tetralogia de Fallot", "Anomalia de Ebstein", "Coarctação da aorta", "Ventrículo único", "Comunicação atrioventricular", "Dobra epicárdica", "Anomalia de Uhl", "Miocardiopatia hipertrófica", "Miocardiopatia dilatada", "Miocardiopatia restritiva"],
    },
}


# ═══════════════════════════════════════════════════════════════════════
# MAPEAMENTO DICOM SR → CAMPOS DO FORMULÁRIO
# ═══════════════════════════════════════════════════════════════════════

def _mm(v):    return round(v, 1)
def _plain(v): return round(v, 2)
def _pct(v):   return round(v, 1)

MAPA_DICOM = {
    "Patient Weight":       [("DADOS ANTROPOMÉTRICOS", "Peso", _plain)],
    "Patient Height":       [("DADOS ANTROPOMÉTRICOS", "Altura",
                              lambda v: round(v*100,1) if v < 10 else round(v,1))],
    "Body Surface Area":    [("DADOS ANTROPOMÉTRICOS", "Superfície corp.", _plain)],

    "Ascending Aortic Diameter":   [("CÂMARAS ESQUERDAS", "Aorta ascend.", _mm)],
    "Aortic Root Diameter":        [("CÂMARAS ESQUERDAS", "Seio aórtico", _mm)],
    "Cardiovascular Orifice Diameter": [
        ("CÂMARAS ESQUERDAS", "Diâm VSVE", _mm),
        ("AORTA / VSVE",      "Diâm VSVE", _mm)],
    "Left Atrium Antero-posterior Systolic Dimension": [("CÂMARAS ESQUERDAS", "AE - Diâm.", _mm)],
    "Left Atrium Systolic Volume": [("CÂMARAS ESQUERDAS", "AE - Vol. bipl.",
        lambda v: round(v/1000,1) if v > 100 else round(v,1))],
    "Left Atrium Systolic Volume Index": [("CÂMARAS ESQUERDAS", "AE - Vol. bipl. index", _plain)],
    "Interventricular Septum Diastolic Thickness": [("CÂMARAS ESQUERDAS", "Septo", _mm)],
    "Left Ventricle Posterior Wall Diastolic Thickness": [("CÂMARAS ESQUERDAS", "Parede post.", _mm)],
    "Left Ventricle Internal End Diastolic Dimension": [("CÂMARAS ESQUERDAS", "DdVE", _mm)],
    "LVIDd": [("CÂMARAS ESQUERDAS", "DdVE", _mm)],
    "Left Ventricle Internal Systolic Dimension": [("CÂMARAS ESQUERDAS", "DsVE", _mm)],
    "LVIDs": [("CÂMARAS ESQUERDAS", "DsVE", _mm)],
    "Left Ventricular Ejection Fraction": [("CÂMARAS ESQUERDAS", "FEVE (Simpson)", _pct)],
    "Relative Wall Thickness": [("CÂMARAS ESQUERDAS", "ERP", _plain)],
    "Left Ventricle Mass":         [("CÂMARAS ESQUERDAS", "Massa VE", _plain)],
    "Left Ventricle Mass by M-mode":[("CÂMARAS ESQUERDAS", "Massa VE", _plain)],

    "Right Atrium Systolic Area": [("CÂMARAS DIREITAS", "AD - Área",
        lambda v: round(v/100,1) if v > 100 else round(v,1))],
    "Right Atrium Systolic Volume": [("CÂMARAS DIREITAS", "AD - Vol.",
        lambda v: round(v/1000,1) if v > 100 else round(v,1))],
    "Right Atrium Systolic Volume Index": [("CÂMARAS DIREITAS", "AD - Vol. index", _plain)],
    "Right Atrium Systolic Pressure": [("CÂMARAS DIREITAS", "AD - PSAP", _plain)],
    "Right Ventricle Basal Diameter": [("CÂMARAS DIREITAS", "VD - Diâm. basal", _mm)],
    "Tricuspid Annular Plane Systolic Excursion": [
        ("CÂMARAS DIREITAS", "VD - TAPSE", _mm)],
    "Right Ventricle S Velocity": [("CÂMARAS DIREITAS", "VD - Onda S",
        lambda v: round(v/10,1) if v > 10 else round(v,1))],
    "Right Ventricular Peak Systolic Pressure": [("CÂMARAS DIREITAS", "VD - PSAP", _plain)],
    "Right Ventricle Outflow Tract Distal Diameter": [
        ("CÂMARAS DIREITAS",      "AP - Diâm.", _mm),
        ("TRICÚSPIDE / PULMONAR", "AP - Diâm.", _mm)],

    "Mitral Valve E-Wave Peak Velocity": [("VALVA MITRAL", "Vel. onda E",
        lambda v: round(v/1000,2) if v > 10 else round(v,2))],
    "Mitral Valve A-Wave Peak Velocity": [("VALVA MITRAL", "Vel. onda A",
        lambda v: round(v/1000,2) if v > 10 else round(v,2))],
    "Mitral Valve E to A Ratio": [("VALVA MITRAL", "Relação E/A", _plain)],
    "Deceleration Time": [("VALVA MITRAL", "Decel. Time", _plain)],
    "Pressure Half-Time": [("AORTA / VSVE", "_pht_ctx", None)],
    "Area by Pressure Half-Time": [("VALVA MITRAL", "Área (PHT)",
        lambda v: round(v/100,2) if v > 10 else round(v,2))],
    "Mitral Valve Flow Area": [("VALVA MITRAL", "Área (PISA)",
        lambda v: round(v/100,2) if v > 10 else round(v,2))],
    "Mitral Annular Plane Systolic Excursion": [("VALVA MITRAL", "MAPSE", _mm)],
    "Mitral Valve A-Wave Duration": [("VALVA MITRAL", "Dur. onda A", _plain)],
    "Peak Instantaneous Flow Rate": [("VALVA MITRAL", "Fluxo (PISA)",
        lambda v: round(v/1000,1) if v > 100 else round(v,1))],
    "Volume Flow": [("VALVA MITRAL", "Volume Regurg.",
        lambda v: round(v/1000,1) if v > 100 else round(v,1))],

    "Velocity Time Integral":   [("AORTA / VSVE", "_vti_ctx", _plain)],
    "Mean Velocity":            [("AORTA / VSVE", "_meanvel_ctx", _plain)],
    "Mean Gradient":            [("AORTA / VSVE", "_meangrad_ctx", _plain)],
    "Peak Velocity":            [("AORTA / VSVE", "_peakvel_ctx", _plain)],
    "Peak Gradient":            [("AORTA / VSVE", "_peakgrad_ctx", _plain)],
    "Continuity Equation by Velocity Time Integral": [("AORTA / VSVE", "AVAo (EC-VTI)",
        lambda v: round(v/100,2) if v > 10 else round(v,2))],
    "Continuity Equation by Peak Velocity": [("AORTA / VSVE", "AVAo (EC-Vmax)",
        lambda v: round(v/100,2) if v > 10 else round(v,2))],
    "Aortic Valve Area Indexed To BSA": [("AORTA / VSVE", "AVAo index", _plain)],
    "Aortic Valve Velocity Ratio": [("AORTA / VSVE", "Vel. Ratio", _plain)],
    "Cardiovascular Orifice Area": [("AORTA / VSVE", "_avao_ctx", None)],
    "Deceleration Slope": [("AORTA / VSVE", "_slope_ctx", None)],
    # PHT Ao extraído via _pht_ctx handler acima

    "Left Ventricular Peak Early Diastolic Tissue Velocity": [("TDI", "_e_prime_raw", None)],
    "LV Peak Diastolic Tissue Velocity During Atrial Systole": [("TDI", "_a_prime_raw", None)],
    "Ratio of MV Peak Velocity to LV Peak Tissue Velocity E-Wave": [("TDI", "_e_e_prime_raw", None)],
    "Ratio of MV Peak Velocity to avg LV Peak Tissue Velocity E-Wave": [("TDI", "E/e' MÉDIO", _plain)],
    "Mean Myocardial Velocity of E' sep and E' lat": [("TDI", "E/e' MÉDIO",
        lambda v: round(v/10,1) if v > 10 else round(v,1))],
    "Left Ventricle E to A Tissue Velocity Ratio": [("TDI", "E/A tecidual", _plain)],

    "Pulmonary Vein Systolic Peak Velocity":    [("TDI", "VP sistólica",
        lambda v: round(v/1000, 2) if v > 10 else round(v, 2))],
    "Pulmonary Vein Diastolic Peak Velocity":   [("TDI", "VP diastólica",
        lambda v: round(v/1000, 2) if v > 10 else round(v, 2))],
    "Pulmonary Vein Systolic to Diastolic Ratio": [("TDI", "VP S/D", _plain)],
    "Left Ventricular Isovolumic Relaxation Time": [("TDI", "TRIV",
        lambda v: round(v, 1))],
}

MEDIDAS_OCULTAS = {
    "Left Ventricle MOD Diam", "Left Atrium MOD Diam",
    "Right Atrium MOD Diam", "Value",
    "Left Ventricle diastolic major axis", "Left Ventricle systolic major axis",
    "Left Ventricular Diastolic Area", "Left Ventricular Systolic Area",
    "Left Atrium systolic major axis", "Left Atrium Systolic Area",
    "Right Atrium Systolic Major Axis",
    "Simpson's Disk Number", "End Diastole", "End Systole",
    "Left Atrium to Aortic Root Ratio",
    "Interventricular Septum to Posterior Wall Thickness Ratio",
}


# ═══════════════════════════════════════════════════════════════════════
# DICOM — leitura
# ═══════════════════════════════════════════════════════════════════════

SR_SOP = {
    "1.2.840.10008.5.1.4.1.1.88.11","1.2.840.10008.5.1.4.1.1.88.22",
    "1.2.840.10008.5.1.4.1.1.88.33","1.2.840.10008.5.1.4.1.1.88.34",
    "1.2.840.10008.5.1.4.1.1.88.35","1.2.840.10008.5.1.4.1.1.88.67",
    "1.2.840.10008.5.1.4.1.1.88.68","1.2.840.10008.5.1.4.1.1.88.72",
}

def _conceito(item):
    try:
        c = item.ConceptNameCodeSequence
        return c[0].CodeMeaning if c else None
    except: return None

def _eh_finding_site(item):
    try:
        if str(getattr(item, "RelationshipType", "")).upper() == "HAS CONCEPT MOD":
            nc = _conceito(item)
            if nc == "Finding Site":
                ccs = getattr(item, "ConceptCodeSequence", None)
                if ccs:
                    return ccs[0].CodeMeaning
    except: pass
    return None

def _capturar(item, nome, ctx, site, out):
    try:
        tem_selecao = False
        if hasattr(item, "ContentSequence"):
            for sub in item.ContentSequence:
                if _conceito(sub) == "Selection Status":
                    tem_selecao = True; break
        for mv in item.MeasuredValueSequence:
            v = getattr(mv,"NumericValue",None)
            u = ""
            try: u = mv.MeasurementUnitsCodeSequence[0].CodeMeaning
            except: pass
            if v is not None:
                out.append({"nome": nome or "Value", "valor": float(v),
                            "unidade": u, "contexto": ctx, "site": site or "", "selecao": tem_selecao})
    except: pass

def extrair_raw(ds, out=None, ctx="", site_herdado=""):
    if out is None: out = []
    for elem in ds:
        if elem.VR != "SQ": continue
        site_deste_nivel = site_herdado
        label_deste_nivel = None
        for item in elem.value:
            site_encontrado = _eh_finding_site(item)
            if site_encontrado: site_deste_nivel = site_encontrado
            nc = _conceito(item)
            if getattr(item, "ValueType", "") == "TEXT" and nc == "Label":
                try: label_deste_nivel = str(item.TextValue)
                except: pass
        for item in elem.value:
            if _eh_finding_site(item): continue
            nc = _conceito(item)
            if getattr(item, "ValueType", "") == "NUM" and nc == "Value" and label_deste_nivel:
                nc = label_deste_nivel
            nctx = f"{ctx} > {nc}" if nc else ctx
            if hasattr(item, "MeasuredValueSequence"):
                _capturar(item, nc, nctx, site_deste_nivel, out)
            else:
                extrair_raw(item, out, nctx, site_deste_nivel)
    return out

def info_paciente(ds):
    campos = {"PatientName":"Nome","PatientID":"ID","AccessionNumber":"Accession",
              "PatientSex":"Sexo","PatientBirthDate":"Nascimento","StudyDate":"Data do Exame",
              "InstitutionName":"Instituição","StudyDescription":"Descrição"}
    r = {}
    for a,l in campos.items():
        try:
            v = getattr(ds,a,None)
            if v:
                v = str(v).strip()
                if a in ("PatientBirthDate", "StudyDate") and len(v) == 8 and v.isdigit():
                    v = f"{v[6:8]}/{v[4:6]}/{v[0:4]}"
                r[l] = v
        except: pass
    return r


# ═══════════════════════════════════════════════════════════════════════
# MAPEAMENTO SR → FORMULÁRIO
# ═══════════════════════════════════════════════════════════════════════

def _media_vals(vals):
    return round(sum(vals)/len(vals), 2) if vals else None

def _ctx_tem(site, *palavras):
    site_low = (site or "").lower()
    return all(p.lower() in site_low for p in palavras)

def mapear_para_form(medidas_raw):
    grupos = {}
    for m in medidas_raw:
        grupos.setdefault(m["nome"], []).append(m)

    resultado = {}

    for nome_dicom, itens in grupos.items():
        if nome_dicom in MEDIDAS_OCULTAS or nome_dicom not in MAPA_DICOM:
            continue
        destinos = MAPA_DICOM[nome_dicom]
        itens_sel = [i for i in itens if i.get("selecao")]
        vals_float = [i["valor"] for i in (itens_sel if itens_sel else itens)]

        for (secao, campo, conv_fn) in destinos:
            if campo == "_e_prime_raw":
                if len(vals_float) >= 2:
                    resultado[("TDI","Vel. e' septal")]  = round(min(vals_float)/10, 1)
                    resultado[("TDI","Vel. e' lateral")] = round(max(vals_float)/10, 1)
                elif len(vals_float)==1:
                    resultado[("TDI","Vel. e' septal")] = round(vals_float[0]/10,1)
                continue
            if campo == "_a_prime_raw":
                if len(vals_float) >= 2:
                    resultado[("TDI","Vel. a' septal")]  = round(min(vals_float)/10,1)
                    resultado[("TDI","Vel. a' lateral")] = round(max(vals_float)/10,1)
                continue
            if campo == "_e_e_prime_raw":
                if len(vals_float) >= 2:
                    resultado[("TDI","Rel. E/E' septal")]  = round(max(vals_float),1)
                    resultado[("TDI","Rel. E/E' lateral")] = round(min(vals_float),1)
                elif len(vals_float)==1:
                    resultado[("TDI","Rel. E/E' septal")] = round(vals_float[0],1)
                continue
            if campo == "_slope_ctx":
                # Deceleration Slope: pega apenas de Aortic Valve
                for it in itens:
                    site = it["site"]; v = it["valor"]
                    if _ctx_tem(site,"aortic"):
                        resultado[("AORTA / VSVE","Decel. Slope")] = round(v,1)
                        break  # um valor é suficiente
                continue
            if campo == "_pht_ctx":
                # Pressure Half-Time: Aortic Valve → PHT Ao; Mitral → PHT
                # Para mitral: valores muito baixos (<100ms) são de regurgitação, ignorar
                for it in itens:
                    site = it["site"]; v = it["valor"]
                    if _ctx_tem(site,"aortic"):
                        # PHT aórtico: usar o maior (insuficiência)
                        atual = resultado.get(("AORTA / VSVE","PHT Ao"), 0)
                        if v > atual:
                            resultado[("AORTA / VSVE","PHT Ao")] = round(v,1)
                    elif _ctx_tem(site,"mitral") and v >= 100:
                        # PHT mitral: usar o MENOR valor ≥ 100ms
                        # (PHT da onda E é menor que PHT de regurgitação/estenose severa)
                        atual = resultado.get(("VALVA MITRAL","PHT"), float('inf'))
                        if v < atual:
                            resultado[("VALVA MITRAL","PHT")] = round(v,1)
                continue
            if campo == "_avao_ctx":
                # Cardiovascular Orifice Area: pega apenas de "Aortic Valve" (mm² → cm²)
                for it in itens:
                    site = it["site"]; v = it["valor"]
                    if _ctx_tem(site,"aortic"):
                        val_cm2 = round(v/100, 2)
                        # Usa a maior área (evita pegar áreas pequenas de outros contextos)
                        atual = resultado.get(("AORTA / VSVE","AVAo (EC-VTI)"), 0)
                        if val_cm2 > atual:
                            resultado[("AORTA / VSVE","AVAo (EC-VTI)")] = val_cm2
                continue
            if campo == "_vti_ctx":
                for it in itens:
                    site = it["site"]; v = it["valor"]
                    # VTI em mm → cm (÷10)
                    val_cm = round(v/10, 1)
                    is_aortic  = _ctx_tem(site,"aortic")
                    is_mitral  = _ctx_tem(site,"mitral")
                    is_rv_ap   = _ctx_tem(site,"right ventricle") or _ctx_tem(site,"pulmonary")
                    # VSVE: site "Left Ventricle" sem ser "outflow" explícito
                    # no Philips EPIQ o site do VSVE é "Left Ventricle"
                    is_vsve    = (_ctx_tem(site,"left ventricle") and not is_aortic)
                    if is_aortic:
                        resultado[("AORTA / VSVE","VTI Ao")] = val_cm
                    elif is_vsve:
                        resultado[("AORTA / VSVE","VTI VSVE")] = val_cm
                    elif is_mitral:
                        resultado[("VALVA MITRAL","VTI mit.")] = val_cm
                    elif is_rv_ap:
                        resultado[("TRICÚSPIDE / PULMONAR","VTI AP")] = val_cm
                continue
            if campo in ("_meanvel_ctx","_meangrad_ctx","_peakvel_ctx","_peakgrad_ctx"):
                for it in itens:
                    site = it["site"]; v = it["valor"]
                    is_aortic    = _ctx_tem(site,"aortic")
                    is_mitral    = _ctx_tem(site,"mitral")
                    is_tricuspid = _ctx_tem(site,"tricuspid")
                    is_vsve      = _ctx_tem(site,"left ventricle") and not is_aortic
                    if campo == "_meangrad_ctx":
                        if is_aortic:
                            # Coleta todos os gradientes médios aórticos → usa média
                            resultado.setdefault("_ao_mean_grad_list", []).append(v)
                        elif is_mitral:
                            # Mantém o maior gradiente médio mitral
                            atual = resultado.get(("VALVA MITRAL","Grad. méd."), 0)
                            resultado[("VALVA MITRAL","Grad. méd.")] = max(round(v,1), atual)
                    elif campo == "_peakgrad_ctx":
                        if is_aortic:
                            # Coleta todos os picos aórticos → usa média
                            resultado.setdefault("_ao_peak_grad_list", []).append(v)
                        elif is_mitral:
                            # Mantém o maior gradiente pico mitral
                            atual = resultado.get(("VALVA MITRAL","Grad. máx."), 0)
                            resultado[("VALVA MITRAL","Grad. máx.")] = max(round(v,1), atual)
                        elif is_tricuspid:
                            atual = resultado.get(("TRICÚSPIDE / PULMONAR","PSAP"), 0)
                            resultado[("TRICÚSPIDE / PULMONAR","PSAP")] = max(round(v,1), atual)
                    elif campo == "_peakvel_ctx":
                        val_ms = round(v/1000, 2)
                        if is_aortic:
                            # Coleta todas as velocidades pico aórticas → usa média
                            resultado.setdefault("_ao_peak_vel_list", []).append(v)
                        elif is_tricuspid:
                            resultado[("TRICÚSPIDE / PULMONAR","Vel. RT")] = val_ms
                        elif is_vsve:
                            resultado[("AORTA / VSVE","_peak_vsve")] = val_ms
                    elif campo == "_meanvel_ctx":
                        if is_vsve:
                            resultado[("AORTA / VSVE","_mean_vsve")] = round(v/10, 1)
                continue

            val = _media_vals(vals_float)
            if val is None: continue
            if conv_fn: val = conv_fn(val)
            resultado[(secao, campo)] = val

    # Consolida listas de gradientes/velocidades aórticas
    # O Philips EPIQ duplica medições — pegar valores únicos e usar o MENOR par
    # (corresponde à medição selecionada pelo operador no aparelho)
    ao_peak_vels = sorted(set(resultado.pop("_ao_peak_vel_list", [])))
    ao_peak_grads= sorted(set(resultado.pop("_ao_peak_grad_list", [])))
    ao_mean_grads= sorted(set(resultado.pop("_ao_mean_grad_list", [])))

    if ao_peak_vels:
        # Menor velocidade pico única → corresponde ao menor gradiente
        resultado[("AORTA / VSVE","_peak_ao")] = round(ao_peak_vels[0]/1000, 2)
    if ao_peak_grads:
        resultado[("AORTA / VSVE","Grad. máx.")] = round(ao_peak_grads[0], 1)
    if ao_mean_grads:
        resultado[("AORTA / VSVE","Grad. méd.")] = round(ao_mean_grads[0], 1)

    _calcular_derivados(resultado)
    return resultado


def _calcular_derivados(resultado):
    def g(sec, campo): return resultado.get((sec, campo))

    peso   = g("DADOS ANTROPOMÉTRICOS","Peso")
    altura = g("DADOS ANTROPOMÉTRICOS","Altura")
    bsa    = g("DADOS ANTROPOMÉTRICOS","Superfície corp.")

    if peso and altura and ("DADOS ANTROPOMÉTRICOS","IMC") not in resultado:
        altura_m = altura/100 if altura > 3 else altura
        if altura_m > 0:
            resultado[("DADOS ANTROPOMÉTRICOS","IMC")] = round(peso/(altura_m**2),1)

    ddve = g("CÂMARAS ESQUERDAS","DdVE")
    dsve = g("CÂMARAS ESQUERDAS","DsVE")
    sep  = g("CÂMARAS ESQUERDAS","Septo")
    pp   = g("CÂMARAS ESQUERDAS","Parede post.")

    if ddve and ("CÂMARAS ESQUERDAS","VDVE") not in resultado:
        d = ddve/10; resultado[("CÂMARAS ESQUERDAS","VDVE")] = round((7*d**3)/(2.4+d),1)
    if dsve and ("CÂMARAS ESQUERDAS","VSVE") not in resultado:
        d = dsve/10; resultado[("CÂMARAS ESQUERDAS","VSVE")] = round((7*d**3)/(2.4+d),1)

    vdve = g("CÂMARAS ESQUERDAS","VDVE")
    vsve = g("CÂMARAS ESQUERDAS","VSVE")
    if vdve and vsve and vdve > 0 and ("CÂMARAS ESQUERDAS","FEVE (Teichholz)") not in resultado:
        resultado[("CÂMARAS ESQUERDAS","FEVE (Teichholz)")] = round(((vdve-vsve)/vdve)*100,1)

    if ddve and sep and pp and ("CÂMARAS ESQUERDAS","ERP") not in resultado:
        resultado[("CÂMARAS ESQUERDAS","ERP")] = round((sep+pp)/ddve,2)

    if ddve and sep and pp and ("CÂMARAS ESQUERDAS","Massa VE") not in resultado:
        resultado[("CÂMARAS ESQUERDAS","Massa VE")] = round(
            0.8*(1.04*(((ddve+sep+pp)/10)**3-(ddve/10)**3))+0.6, 1)

    if bsa and bsa > 0:
        for (s,c,src_s,src_c) in [
            ("CÂMARAS ESQUERDAS","Aorta ascend. index","CÂMARAS ESQUERDAS","Aorta ascend."),
            ("CÂMARAS ESQUERDAS","VDVE index","CÂMARAS ESQUERDAS","VDVE"),
            ("CÂMARAS ESQUERDAS","AE - Vol. bipl. index","CÂMARAS ESQUERDAS","AE - Vol. bipl."),
            ("CÂMARAS DIREITAS","AD - Vol. index","CÂMARAS DIREITAS","AD - Vol."),
        ]:
            v = g(src_s,src_c)
            if v and (s,c) not in resultado:
                resultado[(s,c)] = round(v/bsa,1)

        massa = g("CÂMARAS ESQUERDAS","Massa VE")
        if massa and ("CÂMARAS ESQUERDAS","Massa index") not in resultado:
            resultado[("CÂMARAS ESQUERDAS","Massa index")] = round(massa/bsa,1)

        d  = g("AORTA / VSVE","Diâm VSVE")
        vv = g("AORTA / VSVE","VTI VSVE")
        va = g("AORTA / VSVE","VTI Ao")
        avao = g("AORTA / VSVE","AVAo (EC-VTI)")
        if not avao and d and vv and va:
            avao = round((3.1416*((d/10)/2)**2*vv)/va,2)
            resultado[("AORTA / VSVE","AVAo (EC-VTI)")] = avao
        if avao and ("AORTA / VSVE","AVAo index") not in resultado:
            resultado[("AORTA / VSVE","AVAo index")] = round(avao/bsa,2)

    s_tdi = g("TDI","Rel. E/E' septal")
    l_tdi = g("TDI","Rel. E/E' lateral")
    if s_tdi and l_tdi and ("TDI","E/e' MÉDIO") not in resultado:
        resultado[("TDI","E/e' MÉDIO")] = round((s_tdi+l_tdi)/2,1)

    peak_ao = g("AORTA / VSVE","_peak_ao")
    if peak_ao:
        resultado[("AORTA / VSVE","Vel. máx. Ao")] = peak_ao
        if ("AORTA / VSVE","Grad. máx.") not in resultado:
            resultado[("AORTA / VSVE","Grad. máx.")] = round(4*(peak_ao**2),1)


# ═══════════════════════════════════════════════════════════════════════
# FÓRMULAS REATIVAS
# ═══════════════════════════════════════════════════════════════════════

FORMULAS_CALCULADAS = [
    (("DADOS ANTROPOMÉTRICOS","Superfície corp."),
        [("DADOS ANTROPOMÉTRICOS","Peso"),("DADOS ANTROPOMÉTRICOS","Altura")],
        lambda p,a: round(0.007184*(p**0.425)*((a if a>3 else a*100)**0.725),2) if p and a else None),

    (("DADOS ANTROPOMÉTRICOS","IMC"),
        [("DADOS ANTROPOMÉTRICOS","Peso"),("DADOS ANTROPOMÉTRICOS","Altura")],
        lambda p,a: round(p/((a/100 if a>3 else a)**2),1) if a else None),

    (("CÂMARAS ESQUERDAS","VDVE"),
        [("CÂMARAS ESQUERDAS","DdVE")],
        lambda dd: round((7*((dd/10)**3))/(2.4+(dd/10)),1) if dd else None),

    (("CÂMARAS ESQUERDAS","VSVE"),
        [("CÂMARAS ESQUERDAS","DsVE")],
        lambda ds: round((7*((ds/10)**3))/(2.4+(ds/10)),1) if ds else None),

    (("CÂMARAS ESQUERDAS","FEVE (Teichholz)"),
        [("CÂMARAS ESQUERDAS","VDVE"),("CÂMARAS ESQUERDAS","VSVE")],
        lambda vd,vs: round(((vd-vs)/vd)*100,1) if vd and vd>0 else None),

    (("CÂMARAS ESQUERDAS","ERP"),
        [("CÂMARAS ESQUERDAS","Septo"),("CÂMARAS ESQUERDAS","Parede post."),("CÂMARAS ESQUERDAS","DdVE")],
        lambda s,pp,dd: round((s+pp)/dd,2) if dd else None),

    (("CÂMARAS ESQUERDAS","Massa VE"),
        [("CÂMARAS ESQUERDAS","DdVE"),("CÂMARAS ESQUERDAS","Septo"),("CÂMARAS ESQUERDAS","Parede post.")],
        lambda dd,s,pp: round(0.8*(1.04*(((dd+s+pp)/10)**3-(dd/10)**3))+0.6,1) if dd and s and pp else None),

    (("CÂMARAS ESQUERDAS","Massa index"),
        [("CÂMARAS ESQUERDAS","Massa VE"),("DADOS ANTROPOMÉTRICOS","Superfície corp.")],
        lambda m,bsa: round(m/bsa,1) if bsa else None),

    (("CÂMARAS ESQUERDAS","Aorta ascend. index"),
        [("CÂMARAS ESQUERDAS","Aorta ascend."),("DADOS ANTROPOMÉTRICOS","Superfície corp.")],
        lambda ao,bsa: round(ao/bsa,1) if bsa else None),

    (("CÂMARAS ESQUERDAS","VDVE index"),
        [("CÂMARAS ESQUERDAS","VDVE"),("DADOS ANTROPOMÉTRICOS","Superfície corp.")],
        lambda v,bsa: round(v/bsa,1) if bsa else None),

    (("CÂMARAS ESQUERDAS","AE - Vol. bipl. index"),
        [("CÂMARAS ESQUERDAS","AE - Vol. bipl."),("DADOS ANTROPOMÉTRICOS","Superfície corp.")],
        lambda v,bsa: round(v/bsa,1) if bsa else None),

    (("CÂMARAS DIREITAS","AD - Vol. index"),
        [("CÂMARAS DIREITAS","AD - Vol."),("DADOS ANTROPOMÉTRICOS","Superfície corp.")],
        lambda v,bsa: round(v/bsa,1) if bsa else None),

    (("AORTA / VSVE","AVAo (EC-VTI)"),
        [("AORTA / VSVE","Diâm VSVE"),("AORTA / VSVE","VTI VSVE"),("AORTA / VSVE","VTI Ao")],
        lambda d,vv,va: round((3.1416*((d/10)/2)**2*vv)/va,2) if va else None),

    (("AORTA / VSVE","AVAo index"),
        [("AORTA / VSVE","AVAo (EC-VTI)"),("DADOS ANTROPOMÉTRICOS","Superfície corp.")],
        lambda a,bsa: round(a/bsa,2) if bsa else None),

    (("VALVA MITRAL","Área (PHT)"),
        [("VALVA MITRAL","PHT")],
        lambda pht: round(220/pht, 2) if pht and pht > 0 else None),

    (("VALVA MITRAL","Relação E/A"),
        [("VALVA MITRAL","Vel. onda E"),("VALVA MITRAL","Vel. onda A")],
        lambda e,a: round(e/a,2) if a else None),

    (("TDI","Rel. E/E' septal"),
        [("VALVA MITRAL","Vel. onda E"),("TDI","Vel. e' septal")],
        lambda e,ep: round((e*100)/ep,1) if ep else None),

    (("TDI","Rel. E/E' lateral"),
        [("VALVA MITRAL","Vel. onda E"),("TDI","Vel. e' lateral")],
        lambda e,ep: round((e*100)/ep,1) if ep else None),

    (("TDI","E/e' MÉDIO"),
        [("TDI","Rel. E/E' septal"),("TDI","Rel. E/E' lateral")],
        lambda s,l: round((s+l)/2,1)),

    (("TDI","E/A tecidual"),
        [("TDI","Vel. e' septal"),("TDI","Vel. a' septal")],
        lambda e,a: round(e/a,2) if a else None),
]

CAMPOS_2_DECIMAIS = {("CÂMARAS ESQUERDAS","ERP")}


def _fmt(dest, valor_float) -> str:
    if dest in CAMPOS_2_DECIMAIS:
        return f"{valor_float:.2f}".replace(".", ",")
    s = f"{valor_float:.10g}"
    return s.replace(".", ",")


def _to_float(s):
    try:
        return float(str(s).strip().replace(",", ".")) if str(s).strip() != "" else None
    except (ValueError, TypeError):
        return None


def recalcular(valores: dict) -> dict:
    v = dict(valores)
    for _ in range(len(FORMULAS_CALCULADAS) + 1):
        mudou = False
        for (dest, fontes, fn) in FORMULAS_CALCULADAS:
            srcs = [_to_float(v.get(k)) for k in fontes]
            if any(s is None for s in srcs):
                continue
            try:
                novo = fn(*srcs)
            except (ZeroDivisionError, TypeError, ValueError):
                continue
            if novo is None:
                continue
            novo_fmt = _fmt(dest, novo)
            if v.get(dest) != novo_fmt:
                v[dest] = novo_fmt
                mudou = True
        if not mudou:
            break
    return v


# ═══════════════════════════════════════════════════════════════════════
# REFERÊNCIAS
# ═══════════════════════════════════════════════════════════════════════

def _parse_ref(ref_str):
    if not ref_str or ref_str.strip() in ("","-"): return None
    s = ref_str.strip().replace(",",".")
    m = re.match(r'^([\d.]+)\s*-\s*([\d.]+)$', s)
    if m: return ('range', float(m.group(1)), float(m.group(2)))
    m = re.match(r'^<=\s*([\d.]+)$', s)
    if m: return ('max_eq', float(m.group(1)))
    m = re.match(r'^<\s*([\d.]+)$', s)
    if m: return ('max', float(m.group(1)))
    m = re.match(r'^>=\s*([\d.]+)$', s)
    if m: return ('min_eq', float(m.group(1)))
    m = re.match(r'^>\s*([\d.]+)$', s)
    if m: return ('min', float(m.group(1)))
    return None

def _dentro_ref(valor_str, ref_str):
    try: v = float(str(valor_str).strip().replace(",","."))
    except: return None
    p = _parse_ref(ref_str)
    if p is None: return None
    t = p[0]
    if t=='range':  return p[1] <= v <= p[2]
    if t=='max_eq': return v <= p[1]
    if t=='max':    return v < p[1]
    if t=='min_eq': return v >= p[1]
    if t=='min':    return v > p[1]
    return None


# ═══════════════════════════════════════════════════════════════════════
# LAUDO
# ═══════════════════════════════════════════════════════════════════════

def _gv(valores, secao, campo):
    s = str(valores.get((secao, campo), "")).strip()
    try: return float(s.replace(",","."))
    except: return None

def _ref(sexo, secao, campo_nome):
    for sec, campos in FORMULARIO.items():
        if sec == secao:
            for c in campos:
                if c["name"] == campo_nome:
                    return c["ref_mas"] if sexo=="M" else c["ref_fem"]
    return ""

NOMES_SEG_LAUDO = [
    '', 'anterior basal','antero-septal basal','septal basal',
    'inferior basal','infero-lateral basal','antero-lateral basal',
    'anterior medio','antero-septal medio','septal medio',
    'inferior medio','infero-lateral medio','antero-lateral medio',
    'anterior apical','septal apical','inferior apical',
    'lateral apical','apical',
]

SCORE_NOMES = {1:'normal', 2:'hipocinetico', 3:'acinetico', 4:'discinetico'}


def _gerar_motilidade(wmsi_scores):
    """Gera texto de motilidade parietal baseado nos scores do bullseye."""
    sc = {int(k): int(v) for k, v in wmsi_scores.items()}
    wmsi = sum(sc.values()) / 17

    alterados = {2: [], 3: [], 4: []}
    for seg, score in sc.items():
        if score > 1:
            alterados[score].append(NOMES_SEG_LAUDO[seg])

    if wmsi == 1.0:
        motil = "Espessamento sistólico normal em todos os segmentos do VE"
        wmsi_str = f"WMSI = {wmsi:.2f} — Motilidade parietal normal"
        return motil, wmsi_str

    linhas = []
    for score in [4, 3, 2]:
        segs = alterados[score]
        if segs:
            linhas.append(f"Segmento(s) {SCORE_NOMES[score]}(s): {', '.join(segs)}")

    motil = "Alteracao de motilidade parietal: " + " | ".join(linhas)
    return motil, f"WMSI = {wmsi:.2f}"


def gerar_laudo(valores, sexo, estruturado=None, wmsi_scores=None):
    if estruturado is None:
        estruturado = {}
    if wmsi_scores is None:
        wmsi_scores = {str(i): 1 for i in range(1, 18)}

    def get_est(secao, campo):
        if (secao, campo) in estruturado:
            return estruturado[(secao, campo)]
        return sugerir_dropdown(secao, campo, valores, sexo)

    def gv(sec, campo):
        return _gv(valores, sec, campo)

    # ── VENTRÍCULO ESQUERDO ──────────────────────────────────────────
    ve_tam   = get_est("VENTRÍCULO ESQUERDO", "Tamanho da cavidade")
    ve_geom  = get_est("VENTRÍCULO ESQUERDO", "Geometria ventricular")
    ve_fsist = get_est("VENTRÍCULO ESQUERDO", "Função sistólica")
    ve_diast = get_est("VENTRÍCULO ESQUERDO", "Função diastólica")

    ve_dim_txt = "Ventrículo esquerdo (VE) com dimensões normais" if ve_tam == "Normal" \
                 else f"Ventrículo esquerdo (VE) com {ve_tam.lower()}"
    geom_txt   = "Geometria ventricular: normal" if ve_geom == "Normal" \
                 else f"Geometria ventricular: {ve_geom.lower()}"

    # Função sistólica — usa FEVE das medidas
    feve = gv("CÂMARAS ESQUERDAS", "FEVE (Simpson)")
    if feve is None: feve = gv("CÂMARAS ESQUERDAS", "FEVE (Teichholz)")
    feve_str = f" (FEVE {feve:.0f}%)" if feve else ""
    fsist_txt = f"Função sistólica normal do VE{feve_str}" if ve_fsist == "Normal" \
                else f"Função sistólica do VE {ve_fsist.lower()}{feve_str}"

    # Função diastólica — usa medidas de TDI e fluxo mitral
    e_sep   = gv("TDI", "Vel. e' septal")
    e_lat   = gv("TDI", "Vel. e' lateral")
    ee_med  = gv("TDI", "E/e' MÉDIO")
    ea_mit  = gv("VALVA MITRAL", "Relação E/A")
    dt      = gv("VALVA MITRAL", "Decel. Time")
    ae_vol  = gv("CÂMARAS ESQUERDAS", "AE - Vol. bipl. index")

    if ve_diast == "Normal":
        # Monta texto com parâmetros disponíveis
        params = []
        if e_sep is not None: params.append(f"e' septal {e_sep:.0f} cm/s")
        if e_lat is not None: params.append(f"e' lateral {e_lat:.0f} cm/s")
        if ee_med is not None: params.append(f"E/e' médio {ee_med:.1f}")
        if ea_mit is not None: params.append(f"E/A {ea_mit:.2f}")
        params_str = f" ({', '.join(params)})" if params else ""
        fdiast_txt = f"Função diastólica do VE normal{params_str}"
    else:
        grau = ve_diast  # ex: "Disfunção grau I"
        params = []
        if ee_med is not None: params.append(f"E/e' médio {ee_med:.1f}")
        if ea_mit is not None: params.append(f"E/A {ea_mit:.2f}")
        if dt is not None:     params.append(f"DT {dt:.0f} ms")
        if ae_vol is not None: params.append(f"volume AE {ae_vol:.1f} mL/m²")
        params_str = f" ({', '.join(params)})" if params else ""
        fdiast_txt = f"{grau} do VE{params_str}"

    # Motilidade parietal — Wall Motion
    motil_txt, wmsi_txt = _gerar_motilidade(wmsi_scores)

    # ── ÁTRIO ESQUERDO ───────────────────────────────────────────────
    ae_tam = get_est("ÁTRIO ESQUERDO", "Tamanho da cavidade")
    if ae_tam == "Normal":
        ae_txt = "Átrio esquerdo (AE) com volume normal"
    elif ae_vol:
        ae_txt = f"Átrio esquerdo (AE) com {ae_tam.lower()} (índice biplanar {ae_vol:.1f} mL/m²)"
    else:
        ae_txt = f"Átrio esquerdo (AE) com {ae_tam.lower()}"

    # ── VENTRÍCULO DIREITO ───────────────────────────────────────────
    vd_tam   = get_est("VENTRÍCULO DIREITO", "Tamanho da cavidade")
    vd_fsist = get_est("VENTRÍCULO DIREITO", "Função sistólica")
    tapse    = gv("CÂMARAS DIREITAS", "VD - TAPSE")
    onda_s   = gv("CÂMARAS DIREITAS", "VD - Onda S")
    fac      = gv("CÂMARAS DIREITAS", "VD - FAC")

    vd_dim_txt = "Ventrículo direito (VD) com dimensões normais" if vd_tam == "Normal" \
                 else f"Ventrículo direito (VD) com {vd_tam.lower()}"

    params_vd = []
    if tapse is not None: params_vd.append(f"TAPSE {tapse:.0f} mm")
    if onda_s is not None: params_vd.append(f"onda S' {onda_s:.1f} cm/s")
    if fac is not None:   params_vd.append(f"FAC {fac:.0f}%")
    params_vd_str = f" ({', '.join(params_vd)})" if params_vd else ""
    vd_fsist_txt = f"Função sistólica normal do VD{params_vd_str}" if vd_fsist == "Normal" \
                   else f"Função sistólica do VD reduzida{params_vd_str}"

    psap = gv("CÂMARAS DIREITAS", "VD - PSAP") or gv("CÂMARAS DIREITAS", "AD - PSAP") \
           or gv("TRICÚSPIDE / PULMONAR", "PSAP")
    psap_txt = f"PSAP estimada em {psap:.0f} mmHg" if psap is not None else ""

    # ── ÁTRIO DIREITO ────────────────────────────────────────────────
    ad_tam = get_est("ÁTRIO DIREITO", "Tamanho da cavidade")
    ad_vol = gv("CÂMARAS DIREITAS", "AD - Vol. index")
    if ad_tam == "Normal":
        ad_txt = "Átrio direito (AD) com área e volume normais"
    elif ad_vol:
        ad_txt = f"Átrio direito (AD) com {ad_tam.lower()} (índice {ad_vol:.1f} mL/m²)"
    else:
        ad_txt = f"Átrio direito (AD) com {ad_tam.lower()}"

    # ── VALVAS ───────────────────────────────────────────────────────
    valvas_txt = []
    for valva_sec, nome_valva in [
        ("VALVA AORTA","Valva aórtica"), ("VALVA MITRAL","Valva mitral"),
        ("VALVA TRICÚSPIDE","Valva tricúspide"), ("VALVA PULMONAR","Valva pulmonar")]:
        asp = get_est(valva_sec, "Geral")
        est = get_est(valva_sec, "Estenose")
        ins = get_est(valva_sec, "Refluxo")
        asp_str = "com textura, mobilidade e abertura normais dos folhetos" if asp == "Normal" \
                  else f"com {asp.lower()}"
        refluxos = []
        if est != "Ausente": refluxos.append(f"estenose {est.lower()}")
        if ins != "Ausente": refluxos.append(f"refluxo {ins.lower()}")
        ref_str = "Ausência de sinais de refluxo" if not refluxos \
                  else ". ".join(refluxos).capitalize()
        valvas_txt.append(f"{nome_valva} {asp_str}. {ref_str}")

    # ── VASOS DA BASE ────────────────────────────────────────────────
    ao_raiz  = get_est("AORTA", "Raiz da aorta")
    ao_asc   = get_est("AORTA", "Aorta ascendente")
    ap_tronco= get_est("ARTÉRIA PULMONAR", "Tronco da pulmonar")
    seio_ao  = gv("CÂMARAS ESQUERDAS", "Seio aórtico")
    ao_ascv  = gv("CÂMARAS ESQUERDAS", "Aorta ascend.")
    ao_params = []
    if seio_ao:  ao_params.append(f"seio aórtico {seio_ao:.0f} mm")
    if ao_ascv:  ao_params.append(f"aorta ascendente {ao_ascv:.0f} mm")
    ao_params_str = f" ({', '.join(ao_params)})" if ao_params else ""
    ao_txt = f"Aorta ascendente com calibre normal{ao_params_str}. Paredes com textura normal. Fluxo normal" \
             if ao_raiz == "Normal" and ao_asc == "Normal" \
             else f"Aorta com alterações{ao_params_str} (Raiz: {ao_raiz.lower()}, Ascendente: {ao_asc.lower()})"
    ap_txt = "Artéria Pulmonar com calibre normal. Fluxo normal" \
             if ap_tronco == "Normal" else f"Artéria Pulmonar com {ap_tronco.lower()}"

    # ── PERICÁRDIO ───────────────────────────────────────────────────
    peric = get_est("PERICÁRDIO", "Geral")
    peric_txt = "Textura e deslizamento normais do pericárdico" if peric == "Normal" \
                else peric

    # ── CONGÊNITAS ───────────────────────────────────────────────────
    congenita = get_est("CONGÊNITAS", "Geral")
    cong_txt  = "Situs solitus, levocardia. Concordâncias veno-atrial, átrio-ventricular e ventrículo-arterial. Septos íntegros. Canal arterial não visualizado" \
                if congenita == "Ausente" else f"Presença de {congenita}"

    # ── LINHAS OPCIONAIS VD ──────────────────────────────────────────
    vd_extra = [vd_fsist_txt]
    if psap_txt:
        vd_extra.append(psap_txt)

    return [
        "**CÂMARAS ESQUERDAS**",
        ve_dim_txt, geom_txt,
        motil_txt, wmsi_txt,
        fsist_txt, fdiast_txt, "",
        ae_txt, "",
        "**CÂMARAS DIREITAS**",
        vd_dim_txt, *vd_extra, "",
        ad_txt, "",
        "**VALVAS CARDÍACAS**",
        valvas_txt[0], valvas_txt[1], valvas_txt[2], valvas_txt[3], "",
        "**VASOS DA BASE**", ao_txt, ap_txt, "",
        "**PERICÁRDIO**", peric_txt, "",
        "**CONGÊNITAS**", cong_txt, "",
        "**CONCLUSÃO**",
        "- Câmaras cardíacas com dimensões normais",
        "- Funções sistólica e diastólica biventricular normais",
        "- Valvas cardíacas com aspectos morfofuncionais normais",
        "- Ecodopplercardiograma transtorácico normal",
    ]


# ═══════════════════════════════════════════════════════════════════════
# EXPORTAÇÃO
# ═══════════════════════════════════════════════════════════════════════

def gerar_tabela_txt(valores, sexo):
    W=34; WV=8; WU=8; WR=20; TOTAL=W+WV+WU+WR+7
    SEP_H="="*TOTAL; SEP_L="-"*TOTAL
    def fmt(m,v,u,r):
        return f"  {str(m)[:W].ljust(W)} | {str(v)[:WV].rjust(WV)} | {str(u)[:WU].ljust(WU)} | {str(r)[:WR].ljust(WR)}"
    linhas = [SEP_H, fmt("MEDIDA","VALOR","UNIDADE","REFERÊNCIA"), SEP_H]
    for secao, campos in FORMULARIO.items():
        itens = []
        for campo in campos:
            val = str(valores.get((secao,campo["name"]),"")).strip()
            if val and val!="-":
                ref = campo["ref_mas"] if sexo=="M" else campo["ref_fem"]
                itens.append((campo["name"],val,campo["unit"],ref))
        if itens:
            linhas += ["", SEP_L, f"  {secao}", SEP_L]
            for nome,val,unit,ref in itens:
                linhas.append(fmt(nome,val,unit,ref or "-"))
    return "\n".join(linhas)

def exportar_csv_bytes(paciente, valores, sexo, estruturado, wmsi_scores=None):
    W=34; WV=8; WU=8; WR=20; TOTAL=W+WV+WU+WR+7
    SEP_H="="*TOTAL; SEP_L="-"*TOTAL
    def fmt(m,v,u,r):
        return f"  {str(m)[:W].ljust(W)} | {str(v)[:WV].rjust(WV)} | {str(u)[:WU].ljust(WU)} | {str(r)[:WR].ljust(WR)}"
    buf = io.StringIO()
    def w(line=""): buf.write(line+"\r\n")
    w(SEP_H); w("  ECOCARDIOGRAMA - Banco de Dados de Pesquisa"); w(SEP_H)
    for k,v in paciente.items(): w(fmt(k,v,"-","-"))
    w(SEP_L); w()
    w(gerar_tabela_txt(valores, sexo))
    w(); w(SEP_H); w(); w(SEP_H); w("  LAUDO DESCRITIVO"); w(SEP_H); w()
    for linha in gerar_laudo(valores, sexo, estruturado, wmsi_scores or {str(i):1 for i in range(1,18)}):
        w(f"  {linha.replace('**','')}")
    w(); w(SEP_H)
    return buf.getvalue().encode("utf-8-sig")

def exportar_excel_bytes(paciente, valores, sexo, estruturado, wmsi_scores=None):
    thin  = Side(style="thin", color="CCCCCC")
    BRD   = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    H_FILL= PatternFill("solid", fgColor="1F3864")
    H_FONT= Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    S_FILL= PatternFill("solid", fgColor="2E4D8A")
    S_FONT= Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    E_FILL= PatternFill("solid", fgColor="1A4731")   # verde escuro — seção estruturado
    E_FONT= Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    Z_FILL= PatternFill("solid", fgColor="EBF3FB")
    G_FILL= PatternFill("solid", fgColor="E8F5E9")   # verde claro — linhas do estruturado
    C_FILL= PatternFill("solid", fgColor="FFF2CC")
    D_FONT= Font(name="Calibri", size=10)

    wb = openpyxl.Workbook()

    # ── Aba 1: Ecocardiograma (medidas detalhadas) ────────────────────
    ws = wb.active; ws.title = "Ecocardiograma"

    # Dados do paciente
    row = 1
    for k, v in paciente.items():
        ws.cell(row=row, column=1, value=k).font = Font(name="Calibri", bold=True, size=9, color="1F3864")
        ws.cell(row=row, column=2, value=v).font  = Font(name="Calibri", size=9)
        row += 1
    row += 1

    # Cabeçalho das medidas
    for c, (h, w2) in enumerate(zip(
            ["Seção", "Medida", "Valor", "Unidade", "Ref. Masc.", "Ref. Fem.", "Calc."],
            [26, 26, 10, 10, 18, 18, 8]), 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = CTR; cell.border = BRD
        ws.column_dimensions[get_column_letter(c)].width = w2
    row += 1

    sec_ant = ""; zebra = False
    for secao, campos in FORMULARIO.items():
        if secao != sec_ant:
            for c in range(1, 8):
                cell = ws.cell(row=row, column=c, value=secao if c==1 else "")
                cell.font = S_FONT; cell.fill = S_FILL; cell.border = BRD
                cell.alignment = LEFT if c==1 else CTR
            row += 1; sec_ant = secao; zebra = False
        for campo in campos:
            val  = str(valores.get((secao, campo["name"]), "")).strip()
            ref  = campo["ref_mas"] if sexo=="M" else campo["ref_fem"]
            dentro = _dentro_ref(val, ref) if val else None
            fill = C_FILL if campo["calc"] else (Z_FILL if zebra else None)
            for c, v in enumerate(["", campo["name"], val, campo["unit"],
                                    campo["ref_mas"], campo["ref_fem"],
                                    "⚙" if campo["calc"] else ""], 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = D_FONT; cell.border = BRD
                cell.alignment = CTR if c in (3,4,5,6,7) else LEFT
                if fill: cell.fill = fill
                # Destaca valor fora da referência em vermelho
                if c == 3 and dentro is False:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="C0392B")
            row += 1; zebra = not zebra

    # Seção estruturado logo abaixo das medidas
    row += 1
    for c, (h, w2) in enumerate(zip(
            ["Seção", "Campo", "Valor Selecionado", "", "", "", ""],
            [26, 26, 30, 10, 18, 18, 8]), 1):
        if h:
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = E_FONT; cell.fill = E_FILL; cell.alignment = CTR; cell.border = BRD
    row += 1

    sec_ant = ""; zebra = False
    for secao, itens in ESTRUTURA_DROPDOWNS.items():
        if secao != sec_ant:
            for c in range(1, 4):
                cell = ws.cell(row=row, column=c, value=secao if c==1 else "")
                cell.font = E_FONT; cell.fill = E_FILL; cell.border = BRD
                cell.alignment = LEFT if c==1 else CTR
            row += 1; sec_ant = secao; zebra = False
        for nome in itens:
            val = str(estruturado.get((secao, nome), ""))
            fill = G_FILL if zebra else None
            for c, v in enumerate(["", nome, val], 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = D_FONT; cell.border = BRD
                cell.alignment = LEFT if c <= 2 else CTR
                if fill: cell.fill = fill
            row += 1; zebra = not zebra

    ws.freeze_panes = f"A{len(paciente)+3}"

    # ── Aba 2: Banco de Dados (linha por paciente) ────────────────────
    ws2 = wb.create_sheet("Banco de Dados")

    NOMES_SEG_XLS = [
        '', 'Ant basal','Ant-sep basal','Sep basal',
        'Inf basal','Inf-lat basal','Ant-lat basal',
        'Ant med','Ant-sep med','Sep med',
        'Inf med','Inf-lat med','Ant-lat med',
        'Ant apex','Sep apex','Inf apex','Lat apex','Apex',
    ]
    sc_bd = {int(k): int(v) for k, v in (wmsi_scores or {str(i):1 for i in range(1,18)}).items()}
    wmsi_val_bd = sum(sc_bd.values()) / 17

    info_keys  = list(paciente.keys())
    med_campos = [(sec, c["name"]) for sec, campos in FORMULARIO.items() for c in campos]
    est_campos = [(sec, nome) for sec, itens in ESTRUTURA_DROPDOWNS.items() for nome in itens]
    wmsi_campos = [f"WM {NOMES_SEG_XLS[i]}" for i in range(1, 18)] + ["WMSI"]

    W_FILL = PatternFill("solid", fgColor="FF1A4731")   # verde escuro — cabeçalho WMSI
    W_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    # Cabeçalhos
    col = 1
    for k in info_keys:
        cell = ws2.cell(row=1, column=col, value=k)
        cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = CTR; cell.border = BRD
        ws2.column_dimensions[get_column_letter(col)].width = 16
        col += 1
    for sec, nome in med_campos:
        cell = ws2.cell(row=1, column=col, value=f"{nome}\n({sec})")
        cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = CTR; cell.border = BRD
        ws2.column_dimensions[get_column_letter(col)].width = 14
        col += 1
    for sec, nome in est_campos:
        cell = ws2.cell(row=1, column=col, value=f"{nome}\n({sec})")
        cell.font = E_FONT; cell.fill = E_FILL; cell.alignment = CTR; cell.border = BRD
        ws2.column_dimensions[get_column_letter(col)].width = 18
        col += 1
    for nome in wmsi_campos:
        cell = ws2.cell(row=1, column=col, value=nome)
        cell.font = W_FONT; cell.fill = W_FILL; cell.alignment = CTR; cell.border = BRD
        ws2.column_dimensions[get_column_letter(col)].width = 14
        col += 1

    # Dados
    col = 1
    for k in info_keys:
        ws2.cell(row=2, column=col, value=paciente.get(k, "")).font = D_FONT
        col += 1
    for sec, nome in med_campos:
        ws2.cell(row=2, column=col, value=str(valores.get((sec, nome), ""))).font = D_FONT
        col += 1
    for sec, nome in est_campos:
        ws2.cell(row=2, column=col, value=str(estruturado.get((sec, nome), ""))).font = D_FONT
        col += 1
    for i in range(1, 18):
        ws2.cell(row=2, column=col, value=sc_bd.get(i, 1)).font = D_FONT
        col += 1
    ws2.cell(row=2, column=col, value=round(wmsi_val_bd, 2)).font = D_FONT

    ws2.row_dimensions[1].height = 36
    ws2.freeze_panes = "A2"

    # ── Aba 3: Wall Motion ────────────────────────────────────────────
    if wmsi_scores:
        ws3 = wb.create_sheet("Wall Motion")
        sc  = {int(k): int(v) for k, v in wmsi_scores.items()}
        wmsi_val = sum(sc.values()) / 17

        NOMES_SEG_XLS = [
            '', 'Ant basal','Ant-sep basal','Sep basal',
            'Inf basal','Inf-lat basal','Ant-lat basal',
            'Ant med','Ant-sep med','Sep med',
            'Inf med','Inf-lat med','Ant-lat med',
            'Ant apex','Sep apex','Inf apex','Lat apex','Apex',
        ]
        SCORE_LABEL = {1:'Normal',2:'Hipocinético',3:'Acinético',4:'Discinético'}
        SCORE_COLOR = {
            1:'FF00D000', 2:'FFFFD000',
            3:'FF0099FF', 4:'FFFF3030',
        }
        W3_FILL = {k: PatternFill("solid", fgColor=v) for k,v in SCORE_COLOR.items()}

        # Cabeçalho
        for c, h in enumerate(["Segmento","Score","Classificação"], 1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.font = H_FONT; cell.fill = H_FILL
            cell.alignment = CTR; cell.border = BRD
        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 8
        ws3.column_dimensions["C"].width = 18

        # Dados dos 17 segmentos
        for seg in range(1, 18):
            score = sc.get(seg, 1)
            row   = seg + 1
            fill  = W3_FILL[score]
            for c, val in enumerate([NOMES_SEG_XLS[seg], score, SCORE_LABEL[score]], 1):
                cell = ws3.cell(row=row, column=c, value=val)
                cell.font = D_FONT; cell.border = BRD; cell.fill = fill
                cell.alignment = CTR if c > 1 else LEFT

        # WMSI final
        ws3.cell(row=19, column=1, value="WMSI").font = Font(name="Calibri", bold=True, size=11)
        cell_wmsi = ws3.cell(row=19, column=2, value=round(wmsi_val, 2))
        cell_wmsi.font = Font(name="Calibri", bold=True, size=11)
        cell_wmsi.alignment = CTR

        # Interpretação
        if wmsi_val == 1.0:   interp = "Motilidade parietal normal"
        elif wmsi_val <= 1.5: interp = "Disfunção leve"
        elif wmsi_val <= 2.0: interp = "Disfunção moderada"
        else:                  interp = "Disfunção importante"
        ws3.cell(row=19, column=3, value=interp).font = Font(name="Calibri", bold=True, size=11)
        ws3.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()



# SESSION STATE
# ═══════════════════════════════════════════════════════════════════════

MEDICOS = ["Thiago Gabriel", "Ricardo Lima", "Outro"]

def _init_state():
    if "valores"       not in st.session_state: st.session_state.valores       = {}
    if "paciente"      not in st.session_state: st.session_state.paciente      = {}
    if "sexo"          not in st.session_state: st.session_state.sexo          = "F"
    if "medico_sel"    not in st.session_state: st.session_state.medico_sel    = MEDICOS[0]
    if "medico_outro"  not in st.session_state: st.session_state.medico_outro  = ""


# ═══════════════════════════════════════════════════════════════════════
# SUGESTÃO AUTOMÁTICA DE DROPDOWNS
# ═══════════════════════════════════════════════════════════════════════

def sugerir_dropdown(secao, nome, valores, sexo):
    def av(sec, campo):
        v = _gv(valores, sec, campo)
        r = _ref(sexo, sec, campo)
        ok = _dentro_ref(str(v), r) if v is not None else None
        return v, ok

    if secao == "VENTRÍCULO ESQUERDO":
        if nome == "Tamanho da cavidade":
            ddve, ok = av("CÂMARAS ESQUERDAS","DdVE")
            if ddve is None or ok is True: return "Normal"
            if sexo=="M":
                if ddve<=63: return "Dilatação leve"
                elif ddve<=68: return "Dilatação moderada"
                else: return "Dilatação importante"
            else:
                if ddve<=56: return "Dilatação leve"
                elif ddve<=60: return "Dilatação moderada"
                else: return "Dilatação importante"
        elif nome == "Geometria ventricular":
            erp,  _ = av("CÂMARAS ESQUERDAS","ERP")
            massa,_ = av("CÂMARAS ESQUERDAS","Massa index")
            # Garante float mesmo com vírgula
            if erp is not None:
                try: erp = float(str(erp).replace(",","."))
                except: erp = None
            # Limiares IMVE: 95 g/m² (F) / 115 g/m² (M)
            lim_imve = 115 if sexo == "M" else 95
            if erp is not None and massa is not None:
                erp_alto   = erp > 0.42
                massa_alta = massa > lim_imve
                if not erp_alto and not massa_alta: return "Normal"
                elif erp_alto  and not massa_alta:  return "Remodelamento concêntrico"
                elif not erp_alto and massa_alta:   return "Hipertrofia excêntrica"
                else:                               return "Hipertrofia concêntrica"
            elif erp is not None:
                return "Remodelamento concêntrico" if erp > 0.42 else "Normal"
            elif massa is not None:
                return "Hipertrofia excêntrica" if massa > lim_imve else "Normal"
            return "Normal"
        elif nome == "Função sistólica":
            fe = _gv(valores,"CÂMARAS ESQUERDAS","FEVE (Simpson)")
            if fe is None: fe = _gv(valores,"CÂMARAS ESQUERDAS","FEVE (Teichholz)")
            if fe is None: return "Normal"
            lim = 52 if sexo=="M" else 54
            if fe>=lim: return "Normal"
            elif fe>=41: return "Reduzida de grau leve"
            elif fe>=30: return "Reduzida de grau moderado"
            else: return "Reduzida de grau importante"

    elif secao == "VENTRÍCULO DIREITO":
        if nome == "Tamanho da cavidade":
            vd, ok = av("CÂMARAS DIREITAS","VD - Diâm. basal")
            if vd is None or ok is True: return "Normal"
            if vd<=45: return "Dilatação leve"
            elif vd<=50: return "Dilatação moderada"
            else: return "Dilatação importante"
        elif nome == "Função sistólica":
            tapse,_ = av("CÂMARAS DIREITAS","VD - TAPSE")
            onda_s,_= av("CÂMARAS DIREITAS","VD - Onda S")
            fac,_   = av("CÂMARAS DIREITAS","VD - FAC")
            if ((tapse is not None and tapse<17) or
                (onda_s is not None and onda_s<9.5) or
                (fac is not None and fac<35)):
                return "Reduzida"
            return "Normal"

    elif secao == "ÁTRIO ESQUERDO":
        if nome == "Tamanho da cavidade":
            # Classificação pelo volume biplanar indexado (mL/m²) — ASE 2015
            vol_idx,_ = av("CÂMARAS ESQUERDAS","AE - Vol. bipl. index")
            if vol_idx is not None:
                if vol_idx <= 34: return "Normal"
                elif vol_idx <= 41: return "Dilatação leve"
                elif vol_idx <= 48: return "Dilatação moderada"
                else:              return "Dilatação importante"
            return "Normal"

    elif secao == "ÁTRIO DIREITO":
        if nome == "Tamanho da cavidade":
            vol,_ = av("CÂMARAS DIREITAS","AD - Vol. index")
            if vol is not None:
                lim = 32 if sexo=="M" else 27
                if vol<=lim: return "Normal"
                elif vol<=lim+6: return "Dilatação leve"
                else: return "Dilatação importante"
            area,_ = av("CÂMARAS DIREITAS","AD - Área")
            if area is not None:
                return "Normal" if area<=18 else "Dilatação leve"
            return "Normal"

    elif secao == "AORTA":
        if nome == "Raiz da aorta":
            raiz, ok = av("CÂMARAS ESQUERDAS","Seio aórtico")
            return "Dilatação" if raiz is not None and ok is False else "Normal"
        elif nome == "Aorta ascendente":
            asc, ok = av("CÂMARAS ESQUERDAS","Aorta ascend.")
            return "Dilatação" if asc is not None and ok is False else "Normal"

    elif secao == "ARTÉRIA PULMONAR":
        if nome == "Tronco da pulmonar":
            ap,_ = av("TRICÚSPIDE / PULMONAR","AP - Diâm.")
            return "Dilatação" if ap is not None and ap>25 else "Normal"

    # Padrões
    if nome in ("Estenose","Insuficiência") or secao == "CONGÊNITAS":
        return "Ausente"
    return "Normal"


# ═══════════════════════════════════════════════════════════════════════
# INTERFACE
# ═══════════════════════════════════════════════════════════════════════

def main():
    st.set_page_config(page_title="Eco SR Reader", page_icon="🫀", layout="wide")
    _init_state()

    st.markdown("""
    <style>
    .sec-header {
        background:#2E4D8A; color:white; font-weight:bold; font-size:13px;
        padding:6px 12px; border-radius:4px; margin:12px 0 4px 0;
    }
    .ref-ok  { color:#a6e3a1; font-size:12px; }
    .ref-bad { color:#f38ba8; font-size:12px; font-weight:bold; }
    .ref-neu { color:#89dceb; font-size:12px; }
    .calc-label { color:#f9e2af; font-size:11px; }
    </style>
    """, unsafe_allow_html=True)

    col_title, col_sexo = st.columns([6,1])
    with col_title:
        st.title("🫀 Eco SR Reader")
        st.caption("Formulário Estruturado de Ecocardiograma")
    with col_sexo:
        st.session_state.sexo = st.radio(
            "Sexo (referência)", ["M","F"],
            index=0 if st.session_state.sexo=="M" else 1,
            horizontal=True)

    st.divider()

    # ── Upload DICOM SR ──────────────────────────────────────────────
    with st.expander("📂 Carregar DICOM SR", expanded=True):
        uploaded = st.file_uploader(
            "Envie um ou mais arquivos DICOM SR",
            type=None, accept_multiple_files=True,
            help="Selecione os arquivos DICOM SR do ecocardiograma diretamente pelo navegador.")

        if uploaded:
            srs_validos = []
            diagnosticos = []
            for f in uploaded:
                raw_bytes = f.read()
                modality = sop = "—"
                tem_content = aceito = False
                erro_msg = None
                try:
                    import io
                    ds_test = pydicom.dcmread(io.BytesIO(raw_bytes), force=True, stop_before_pixels=True)
                    modality = str(getattr(ds_test,"Modality","")).strip().upper()
                    sop      = str(getattr(ds_test,"SOPClassUID","")).strip()
                    tem_content = hasattr(ds_test, "ContentSequence")
                    aceito = (modality == "SR") or (sop in SR_SOP) or tem_content
                    if aceito:
                        srs_validos.append((f.name, raw_bytes))
                except Exception as e:
                    erro_msg = str(e)

                diagnosticos.append(f"**{f.name}**: Modality=`{modality}`, SOP=`{sop}`, Seq=`{tem_content}`" + (f" (Erro: {erro_msg})" if erro_msg else ""))

            with st.expander("🔍 Diagnóstico dos arquivos enviados (Clique para ver)", expanded=not srs_validos):
                for d in diagnosticos:
                    st.write(d)

            if not srs_validos:
                st.warning("Nenhum arquivo SR válido encontrado.")
            else:
                nomes = [n for n,_ in srs_validos]
                idx = 0
                if len(srs_validos)>1:
                    sel = st.selectbox("Selecione o SR:", nomes)
                    idx = nomes.index(sel)
                else:
                    st.info(f"Arquivo SR detectado: **{nomes[0]}**")

                nome_sel, bytes_sel = srs_validos[idx]

                if st.button("✅ Carregar SR selecionado", type="primary"):
                    with st.spinner("Lendo DICOM SR..."):
                        ds = pydicom.dcmread(io.BytesIO(bytes_sel), force=True)
                        st.session_state.paciente = info_paciente(ds)
                        raw = extrair_raw(ds)
                        mapeado = mapear_para_form(raw)
                        novos = {}
                        for (sec,campo), val in mapeado.items():
                            if campo.startswith("_"): continue
                            dest = (sec,campo)
                            novos[dest] = _fmt(dest, float(val))
                        novos_calc = recalcular(novos)
                        st.session_state.valores = novos_calc
                        for (sec,campo), val in novos_calc.items():
                            st.session_state[f"inp_{sec}_{campo}"] = val
                        sx = st.session_state.paciente.get("Sexo","")
                        if sx.upper() in ("M","MALE","MASCULINO"): st.session_state.sexo="M"
                        elif sx.upper() in ("F","FEMALE","FEMININO"): st.session_state.sexo="F"
                    preenchidos = sum(1 for v in st.session_state.valores.values() if v)
                    st.success(f"✅ {preenchidos} campos preenchidos a partir de {nome_sel}")
                    st.rerun()

    # ── Cálculo de valores ───────────────────────────────────────────
    valores_str = st.session_state.valores

    def get_float(sec, campo):
        return _to_float(valores_str.get((sec,campo),""))

    valores_calc = {}
    for (dest, fontes, fn) in FORMULAS_CALCULADAS:
        srcs = [get_float(s,c) for s,c in fontes]
        if any(x is None for x in srcs): continue
        try: novo = fn(*srcs)
        except: continue
        if novo is None: continue
        valores_calc[dest] = _fmt(dest, novo)

    valores_exibir = {**valores_calc, **{k:v for k,v in valores_str.items() if v}}

    for key, val in valores_exibir.items():
        wkey = f"inp_{key[0]}_{key[1]}"
        if wkey not in st.session_state:
            st.session_state[wkey] = val

    # ── Dropdowns estruturados ───────────────────────────────────────
    estruturado_atual = {}
    sexo_atual = st.session_state.sexo
    for secao, itens in ESTRUTURA_DROPDOWNS.items():
        for nome in itens:
            key  = f"estr_{secao}_{nome}"
            skey = f"sug_{secao}_{nome}"
            nova_sug = sugerir_dropdown(secao, nome, valores_exibir, sexo_atual)
            if nova_sug != st.session_state.get(skey):
                st.session_state[key]  = nova_sug
                st.session_state[skey] = nova_sug
            val = st.session_state.get(key) or nova_sug
            st.session_state[key] = val
            estruturado_atual[(secao, nome)] = val

    # ── Tabs ─────────────────────────────────────────────────────────
    tab_pac, tab_med, tab_est, tab_wmsi, tab_lau, tab_estresse = st.tabs(
        ["👤 Dados do Paciente","📋 Medidas","📊 Estruturado","🫀 Wall Motion","📝 Laudo", "📋 Eco estresse"])

    with tab_pac:
        st.subheader("👤 Dados do Paciente")
        if st.session_state.paciente:
            for k,v in st.session_state.paciente.items():
                novo = st.text_input(k, value=v, key=f"pac_{k}")
                if novo != v: st.session_state.paciente[k] = novo
        else:
            st.info("Nenhum dado de paciente carregado. Carregue um SR ou preencha manualmente.")

        st.markdown("---")
        st.markdown("**Médico Responsável**")
        sel = st.selectbox("Selecionar médico", MEDICOS,
                           index=MEDICOS.index(st.session_state.medico_sel)
                                 if st.session_state.medico_sel in MEDICOS else 0,
                           key="medico_sel_widget", label_visibility="collapsed")
        st.session_state.medico_sel = sel
        if sel == "Outro":
            outro = st.text_input("Nome do médico", value=st.session_state.medico_outro,
                                  key="medico_outro_widget")
            st.session_state.medico_outro = outro
            medico_final = outro.strip() or "Não informado"
        else:
            medico_final = sel
        st.session_state.paciente["Médico Responsável"] = medico_final

    with tab_med:
        st.subheader("📋 Formulário de Medidas")
        sexo = st.session_state.sexo
        valores_editados = {}

        for secao, campos in FORMULARIO.items():
            st.markdown(f'<div class="sec-header">{secao}</div>', unsafe_allow_html=True)
            h1,h2,h3,h4,_ = st.columns([3,1.2,0.7,1.8,0.3])
            h1.markdown("**Medida**"); h2.markdown("**Valor**")
            h3.markdown("**Un.**");    h4.markdown("**Referência**")

            for campo in campos:
                key  = (secao, campo["name"])
                wkey = f"inp_{secao}_{campo['name']}"
                ref  = campo["ref_mas"] if sexo=="M" else campo["ref_fem"]
                is_calc = campo["calc"]

                val_calc   = valores_calc.get(key,"")
                val_manual = valores_str.get(key,"")
                val_widget = st.session_state.get(wkey,"")
                if val_calc and val_widget==val_manual and val_widget!=val_calc:
                    st.session_state[wkey] = val_calc

                val_exibir = valores_exibir.get(key,"")
                c1,c2,c3,c4,c5 = st.columns([3,1.2,0.7,1.8,0.3])

                with c1:
                    if is_calc:
                        st.markdown(f'<span class="calc-label">⚙ {campo["name"]}</span>', unsafe_allow_html=True)
                    else:
                        st.markdown(f"<span style='font-size:13px'>{campo['name']}</span>", unsafe_allow_html=True)
                with c2:
                    novo_val = st.text_input(campo["name"], key=wkey, label_visibility="collapsed")
                    valores_editados[key] = novo_val
                with c3:
                    st.markdown(f"<span style='color:#6c7086;font-size:12px'>{campo['unit']}</span>", unsafe_allow_html=True)
                with c4:
                    dentro = _dentro_ref(val_exibir, ref) if val_exibir else None
                    if dentro is True:
                        st.markdown(f'<span class="ref-ok">✓ {ref}</span>', unsafe_allow_html=True)
                    elif dentro is False:
                        st.markdown(f'<span class="ref-bad">⚠ {ref}</span>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<span class="ref-neu">{ref}</span>', unsafe_allow_html=True)
                with c5:
                    if is_calc:
                        st.markdown('<span class="calc-label">⚙</span>', unsafe_allow_html=True)

        st.session_state.valores = recalcular(valores_editados)

    with tab_est:
        st.subheader("📊 Estruturado")
        for secao, itens in ESTRUTURA_DROPDOWNS.items():
            st.markdown(f'<div class="sec-header">{secao}</div>', unsafe_allow_html=True)
            for nome, opcoes in itens.items():
                st.selectbox(nome, opcoes, key=f"estr_{secao}_{nome}")


    with tab_wmsi:
        st.subheader("🫀 Wall Motion Score Index (WMSI)")

        if "wmsi_scores" not in st.session_state:
            st.session_state.wmsi_scores = {str(i): 1 for i in range(1, 18)}


        # Funções de geometria SVG removidas — bullseye retirado
        sc = st.session_state.wmsi_scores
        wmsi_val = sum(int(v) for v in sc.values()) / 17

        # Selectboxes sincronizados
        st.markdown("#### Scores por segmento")
        NOMES_SEG = [
            "Ant BASAL","Ant-sep BASAL","Inf-sep BASAL","Inf BASAL","Inf-lat BASAL","Ant-lat BASAL",
            "Ant MEDIAL","Ant-sep MEDIAL","Inf-sep MEDIAL","Inf MEDIAL","Inf-lat MEDIAL","Ant-lat MEDIAL",
            "Ant APEX","Sep APEX","Inf APEX","Lat APEX","Apical",
        ]
        OPCOES = ["1 - Normal","2 - Hipocinético","3 - Acinético","4 - Discinético"]

        col_bn, col_bh, _ = st.columns([1, 1, 4])
        with col_bn:
            if st.button("✅ Todos: 1 - Normal", use_container_width=True):
                for i in range(1, 18):
                    st.session_state.wmsi_scores[str(i)] = 1
                    st.session_state[f"wmsi_seg_{i}"] = "1 - Normal"
                st.rerun()
        with col_bh:
            if st.button("⚠️ Todos: 2 - Hipocinético", use_container_width=True):
                for i in range(1, 18):
                    st.session_state.wmsi_scores[str(i)] = 2
                    st.session_state[f"wmsi_seg_{i}"] = "2 - Hipocinético"
                st.rerun()

        for i in range(1, 18):
            wkey = f"wmsi_seg_{i}"
            atual = int(sc.get(str(i), 1)) - 1
            col_idx = (i - 1) % 6
            if col_idx == 0:
                cols = st.columns([2,1,2,1,2,1])
            with cols[(col_idx % 3)*2]:
                st.markdown(f"<span style='font-size:12px'>{i}. {NOMES_SEG[i-1]}</span>",
                            unsafe_allow_html=True)
            with cols[(col_idx % 3)*2+1]:
                novo = st.selectbox(f"Segmento {i}", OPCOES, index=atual,
                                    key=wkey, label_visibility="collapsed")
                st.session_state.wmsi_scores[str(i)] = int(novo[0])

        wmsi_val = sum(int(v) for v in st.session_state.wmsi_scores.values()) / 17
        st.markdown(f"### WMSI = {wmsi_val:.2f}")
        if wmsi_val == 1.0:
            st.success("Motilidade normal")
        elif wmsi_val <= 1.5:
            st.warning("Disfunção leve")
        elif wmsi_val <= 2.0:
            st.warning("Disfunção moderada")
        else:
            st.error("Disfunção importante")

    with tab_lau:
        st.subheader("📝 Laudo Descritivo")
        tem_dados = any(v for v in st.session_state.valores.values())
        if not tem_dados:
            st.info("Nenhuma medida preenchida.")
        else:
            tabela_str  = gerar_tabela_txt(valores_exibir, sexo)
            linhas_laudo= gerar_laudo(valores_exibir, sexo, estruturado_atual, st.session_state.wmsi_scores)
            texto_laudo = "\n".join(l.replace("**","") for l in linhas_laudo)
            conteudo    = tabela_str + "\n\n" + "="*75 + "\n  LAUDO DESCRITIVO\n" + "="*75 + "\n\n" + texto_laudo
            st.text_area("Laudo completo (editável)", value=conteudo, height=600, key="laudo_texto")
            nome_pac = st.session_state.paciente.get("Nome","paciente").replace(" ","_").replace("/","-")
            st.download_button("📄 Baixar Laudo (.txt)",
                               data=conteudo.encode("utf-8"),
                               file_name=f"laudo_{nome_pac}.txt",
                               mime="text/plain")

    with tab_estresse:
        st.subheader("📋 Eco Estresse com Dobutamina")

        # ── Inicialização no session_state ──────────────────────────────
        if "eco_estresse" not in st.session_state:
            st.session_state.eco_estresse = {
                "idade": None,
                # Tabela protocolo
                "Dose":     {"Repouso": "Repouso", "0-3": "5",  "3-6": "10", "6-9": "20", "9-12": "30", "12-15": "40", "Recup.": "Recup."},
                "Atropina": {"Repouso": "",         "0-3": "",   "3-6": "",   "6-9": "0,5","9-12": "",   "12-15": "",   "Recup.": ""},
                "PAs":      {"Repouso": "",         "0-3": "",   "3-6": "",   "6-9": "",   "9-12": "",   "12-15": "",   "Recup.": ""},
                "PAd":      {"Repouso": "",         "0-3": "",   "3-6": "",   "6-9": "",   "9-12": "",   "12-15": "",   "Recup.": ""},
                "FC":       {"Repouso": "",         "0-3": "",   "3-6": "",   "6-9": "",   "9-12": "",   "12-15": "",   "Recup.": ""},
                # Campos narrativos
                "dose_inicio": "",
                "dose_pico": "",
                "atrop_total": "",
                "antecedentes": "",
                "medicacoes": "",
                "sintomatologia": "O paciente não referiu sintomas durante a realização do exame",
                "ecg_repouso": "Ritmo sinusal, regular, sem alterações de ST-T.",
                "ecg_pico": "Taquicardia sinusal, sem alterações significativas em relação ao padrão de repouso.",
                "analise_segmentar": "Ausência de alteração na contratilidade segmentar do V.E. no repouso. Observado aumento da mobilidade e espessamento sistólico de todas as paredes do V.E. durante o estresse farmacológico",
                "conclusao": "",
            }

        ee = st.session_state.eco_estresse
        estagios = ["Repouso", "0-3", "3-6", "6-9", "9-12", "12-15", "Recup."]
        doses    = ["Repouso", "5",   "10",  "20",  "30",   "40",    "Recup."]

        # ── Cabeçalho: Idade e FC calculadas ────────────────────────────
        st.markdown("##### Parâmetros do Paciente")
        col_id, col_fcmax, col_fcsub = st.columns(3)
        with col_id:
            idade_val = st.number_input(
                "Idade (anos)", min_value=1, max_value=120,
                value=int(ee["idade"]) if ee["idade"] else 50,
                step=1, key="ee_idade"
            )
            ee["idade"] = idade_val

        fc_max  = 220 - idade_val
        fc_sub  = round(fc_max * 0.85)

        with col_fcmax:
            st.metric("FC máxima prevista (bpm)", fc_max, help="220 − idade")
        with col_fcsub:
            st.metric("FC submáxima prevista (bpm)", fc_sub, help="FC máxima × 0,85")

        st.divider()

        # ── Informações Técnicas ────────────────────────────────────────
        st.markdown("##### Informações Técnicas")
        col_di, col_dp, col_at = st.columns(3)
        with col_di:
            ee["dose_inicio"] = st.text_input(
                "Dose início (mcg/kg/min)", value=ee["dose_inicio"],
                placeholder="ex: 5", key="ee_dose_inicio")
        with col_dp:
            ee["dose_pico"] = st.text_input(
                "Dose pico (mcg/kg/min)", value=ee["dose_pico"],
                placeholder="ex: 40", key="ee_dose_pico")
        with col_at:
            ee["atrop_total"] = st.text_input(
                "Atropina total (mg)", value=ee["atrop_total"],
                placeholder="ex: 1", key="ee_atrop_total")

        st.markdown("##### Informações Pessoais")
        col_ant, col_med = st.columns(2)
        with col_ant:
            ee["antecedentes"] = st.text_input(
                "Antecedentes", value=ee["antecedentes"],
                placeholder="HAS, DM, DAC prévia…", key="ee_antecedentes")
        with col_med:
            ee["medicacoes"] = st.text_input(
                "Medicações em uso", value=ee["medicacoes"],
                key="ee_medicacoes")

        st.markdown("##### Sintomatologia")
        ee["sintomatologia"] = st.text_area(
            "Sintomatologia", value=ee["sintomatologia"], height=68,
            label_visibility="collapsed", key="ee_sintomatologia")

        st.markdown("##### Eletrocardiograma (ECG)")
        col_ecgr, col_ecgp = st.columns(2)
        with col_ecgr:
            ee["ecg_repouso"] = st.text_input(
                "ECG repouso", value=ee["ecg_repouso"], key="ee_ecg_repouso")
        with col_ecgp:
            ee["ecg_pico"] = st.text_input(
                "ECG pico estresse", value=ee["ecg_pico"], key="ee_ecg_pico")

        st.markdown("##### Análise Segmentar")
        ee["analise_segmentar"] = st.text_area(
            "Análise segmentar", value=ee["analise_segmentar"], height=80,
            label_visibility="collapsed", key="ee_analise_seg")

        st.divider()

        # ── Tabela de estágios ──────────────────────────────────────────
        st.markdown("##### Protocolo de Dobutamina")

        # Linha de cabeçalho da tabela (estágios de tempo)
        header_cols = st.columns([2] + [1]*7)
        header_cols[0].markdown("**Parâmetro**")
        header_cols[1].markdown("**Repouso**")
        header_cols[2].markdown("**0 – 3 min**")
        header_cols[3].markdown("**3 – 6 min**")
        header_cols[4].markdown("**6 – 9 min**")
        header_cols[5].markdown("**9 – 12 min**")
        header_cols[6].markdown("**12 – 15 min**")
        header_cols[7].markdown("**Recup.**")

        # Linha: Dose (mcg/kg/min) — somente leitura
        dose_cols = st.columns([2] + [1]*7)
        dose_cols[0].markdown("**Dose** *(mcg/kg/min)*")
        dose_labels = ["Repouso", "5", "10", "20", "30", "40", "Recup."]
        for i, lbl in enumerate(dose_labels):
            dose_cols[i+1].markdown(f"<div style='text-align:center;padding-top:8px'>{lbl}</div>", unsafe_allow_html=True)

        # Linha: Atropina (mg)
        atrop_cols = st.columns([2] + [1]*7)
        atrop_cols[0].markdown("**Atropina** *(mg)*")
        for i, est in enumerate(estagios):
            ee["Atropina"][est] = atrop_cols[i+1].text_input(
                label=f"Atropina_{est}", value=ee["Atropina"][est],
                label_visibility="collapsed", key=f"ee_atrop_{est}"
            )

        # Linha: PAs (mmHg)
        pas_cols = st.columns([2] + [1]*7)
        pas_cols[0].markdown("**PAs** *(mmHg)*")
        for i, est in enumerate(estagios):
            ee["PAs"][est] = pas_cols[i+1].text_input(
                label=f"PAs_{est}", value=ee["PAs"][est],
                label_visibility="collapsed", key=f"ee_PAs_{est}"
            )

        # Linha: PAd (mmHg)
        pad_cols = st.columns([2] + [1]*7)
        pad_cols[0].markdown("**PAd** *(mmHg)*")
        for i, est in enumerate(estagios):
            ee["PAd"][est] = pad_cols[i+1].text_input(
                label=f"PAd_{est}", value=ee["PAd"][est],
                label_visibility="collapsed", key=f"ee_PAd_{est}"
            )

        # Linha: FC (bpm)
        fc_cols = st.columns([2] + [1]*7)
        fc_cols[0].markdown("**FC** *(bpm)*")
        for i, est in enumerate(estagios):
            ee["FC"][est] = fc_cols[i+1].text_input(
                label=f"FC_{est}", value=ee["FC"][est],
                label_visibility="collapsed", key=f"ee_FC_{est}"
            )

        st.divider()

        # ── Conclusão ───────────────────────────────────────────────────
        st.markdown("##### Conclusão")
        conclusao_default = (
            "- Exame (eficaz/submáximo), (não) atingindo 85% da FC máxima calculada para idade;\n"
            "- Ausência de alterações eletrocardiográficas sugestivas de isquemia;\n"
            "- Exame (positivo/negativo) para isquemia miocárdica induzida pelo estresse;\n"
            "- (Ausência de viabilidade miocárdica / Presença de viabilidade miocárdica em região... do ventrículo esquerdo)"
        )
        if not ee["conclusao"]:
            ee["conclusao"] = conclusao_default
        ee["conclusao"] = st.text_area(
            "Conclusão", value=ee["conclusao"], height=120,
            label_visibility="collapsed", key="ee_conclusao")

        st.divider()

        # ── Exportar laudo completo como texto ──────────────────────────
        def gerar_txt_estresse():
            sep  = "=" * 75
            sep2 = "-" * 102
            col_w = 12
            par_w = 18

            dose_i = ee["dose_inicio"] or "__"
            dose_p = ee["dose_pico"]   or "__"
            atrop  = ee["atrop_total"] or "__"

            doc = []
            doc.append(sep)
            doc.append("  ECOCARDIOGRAMA COM ESTRESSE FARMACOLÓGICO")
            doc.append(sep)
            doc.append("")

            # Informações Técnicas
            doc.append("INFORMAÇÕES TÉCNICAS")
            doc.append("")
            doc.append(
                f"Exame realizado no repouso e durante administração endovenosa de Dobutamina, "
                f"iniciada na dose de {dose_i}mcg/kg/min, seguido de aumento gradual da infusão, "
                f"com pico do estresse em {dose_p}mcg/kg/min, após administração de Atropina ({atrop}mg). "
                f"Exame interrompido por devido fim do protocolo. "
                f"Realizado infusão de Metoprolol endovenoso (5mg) na fase de recuperação."
            )
            doc.append("")

            # Informações Pessoais
            doc.append("INFORMAÇÕES PESSOAIS")
            doc.append("")
            doc.append(f"Antecedentes: {ee['antecedentes']}")
            doc.append(f"Medicações em uso: {ee['medicacoes']}")
            doc.append("")

            # Sintomatologia
            doc.append("SINTOMATOLOGIA")
            doc.append("")
            doc.append(ee["sintomatologia"])
            doc.append("")

            # ECG
            doc.append("ELETROCARDIOGRAMA (ECG)")
            doc.append("")
            doc.append(f"ECG repouso: {ee['ecg_repouso']}")
            doc.append(f"ECG pico estresse: {ee['ecg_pico']}")
            doc.append("")

            # Análise Segmentar
            doc.append("ANÁLISE SEGMENTAR")
            doc.append("")
            doc.append(ee["analise_segmentar"])
            doc.append("")

            # Dados do Exame (tabela)
            doc.append("DADOS DO EXAME")
            doc.append(f"Idade: {idade_val} anos")
            doc.append(f"FC máxima prevista: {fc_max} bpm  (220 - idade)")
            doc.append(f"FC submáxima prevista: {fc_sub} bpm  (FC máxima x 0,85)")
            doc.append("")

            header = "Parâmetro".ljust(par_w) + "".join(e.ljust(col_w) for e in estagios)
            doc.append(header)
            doc.append("-" * len(header))
            doc.append("Tempo (min)".ljust(par_w)       + "".join(e.ljust(col_w) for e in ["0", "0-3", "3-6", "6-9", "9-12", "12-15", ">15"]))
            doc.append("Dose (mcg/kg/min)".ljust(par_w) + "".join(d.ljust(col_w) for d in dose_labels))
            for param in ["Atropina", "PAs", "PAd", "FC"]:
                unidade = {"Atropina": "(mg)", "PAs": "(mmHg)", "PAd": "(mmHg)", "FC": "(bpm)"}[param]
                row = f"{param} {unidade}".ljust(par_w)
                row += "".join((ee[param][e] or "").ljust(col_w) for e in estagios)
                doc.append(row)
            doc.append("")

            # Conclusão
            doc.append("CONCLUSÃO")
            doc.append("")
            doc.append(ee["conclusao"])
            doc.append("")

            return "\n".join(doc)

        nome_pac_ee = st.session_state.paciente.get("Nome", "paciente").replace(" ", "_").replace("/", "-")

        txt_preview = gerar_txt_estresse()
        #st.text_area("Pré-visualização do laudo", value=txt_preview, height=300, key="ee_preview")

        st.download_button(
            "📄 Baixar Eco Estresse (.txt)",
            data=txt_preview.encode("utf-8"),
            file_name=f"eco_estresse_{nome_pac_ee}.txt",
            mime="text/plain",
        )

        # ── Gráfico de linhas: PAs, PAd e FC ───────────────────────────
        st.divider()
        st.markdown("##### 📈 Gráfico de Tendência")

        import matplotlib.pyplot as plt
        import matplotlib.ticker as ticker

        def parse_num(val):
            """Converte string (aceita vírgula) para float; retorna None se inválido."""
            try:
                return float(str(val).replace(",", ".").strip())
            except (ValueError, TypeError):
                return None

        # Montar séries numéricas por estágio
        x_labels = estagios  # ["Repouso", "0-3", "3-6", "6-9", "9-12", "12-15", "Recup."]
        pas_vals = [parse_num(ee["PAs"][e])  for e in x_labels]
        pad_vals = [parse_num(ee["PAd"][e])  for e in x_labels]
        fc_vals  = [parse_num(ee["FC"][e])   for e in x_labels]

        # Só plota se houver ao menos 2 pontos em qualquer série
        tem_grafico = (
            sum(v is not None for v in pas_vals) >= 2 or
            sum(v is not None for v in pad_vals) >= 2 or
            sum(v is not None for v in fc_vals)  >= 2
        )

        if not tem_grafico:
            st.info("Preencha ao menos 2 estágios de PAs, PAd ou FC para visualizar o gráfico.")
        else:
            x_idx = list(range(len(x_labels)))

            fig, ax1 = plt.subplots(figsize=(10, 4.5))
            fig.patch.set_facecolor("#0e1117")
            ax1.set_facecolor("#1a1d2e")

            # Eixo esquerdo: PA
            cor_pas = "#e74c3c"
            cor_pad = "#e67e22"
            cor_fc  = "#3498db"

            def plot_serie(ax, x_idx, vals, label, color, marker="o", linestyle="-"):
                xs = [x_idx[i] for i, v in enumerate(vals) if v is not None]
                ys = [v          for v in vals if v is not None]
                if len(xs) >= 1:
                    ax.plot(xs, ys, color=color, marker=marker, linestyle=linestyle,
                            linewidth=2, markersize=6, label=label)
                    for xi, yi in zip(xs, ys):
                        ax.annotate(f"{yi:.0f}", (xi, yi),
                                    textcoords="offset points", xytext=(0, 8),
                                    ha="center", fontsize=8, color=color)

            plot_serie(ax1, x_idx, pas_vals, "PAs (mmHg)", cor_pas)
            plot_serie(ax1, x_idx, pad_vals, "PAd (mmHg)", cor_pad)

            ax1.set_ylabel("Pressão Arterial (mmHg)", color="white", fontsize=10)
            ax1.tick_params(axis="y", labelcolor="white")
            ax1.tick_params(axis="x", labelcolor="white")
            ax1.set_xticks(x_idx)
            ax1.set_xticklabels(x_labels, fontsize=9, color="white")
            ax1.yaxis.set_minor_locator(ticker.AutoMinorLocator())
            ax1.grid(axis="y", color="#2c2f45", linewidth=0.7, linestyle="--")
            ax1.spines["bottom"].set_color("#444")
            ax1.spines["left"].set_color("#444")
            ax1.spines["top"].set_visible(False)
            ax1.spines["right"].set_visible(False)

            # Eixo direito: FC
            ax2 = ax1.twinx()
            ax2.set_facecolor("#1a1d2e")
            plot_serie(ax2, x_idx, fc_vals, "FC (bpm)", cor_fc, marker="s", linestyle="--")
            ax2.set_ylabel("FC (bpm)", color=cor_fc, fontsize=10)
            ax2.tick_params(axis="y", labelcolor=cor_fc)
            ax2.spines["right"].set_color(cor_fc)
            ax2.spines["top"].set_visible(False)
            ax2.spines["left"].set_visible(False)
            ax2.spines["bottom"].set_color("#444")

            # Linha FC submáxima
            ax2.axhline(y=fc_sub, color=cor_fc, linestyle=":", linewidth=1.2, alpha=0.6)
            ax2.annotate(f"FC submáx {fc_sub} bpm", xy=(x_idx[-1], fc_sub),
                         xytext=(-5, 5), textcoords="offset points",
                         ha="right", fontsize=8, color=cor_fc, alpha=0.8)

            # Legenda unificada
            lines1, labels1 = ax1.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax1.legend(lines1 + lines2, labels1 + labels2,
                       loc="upper left", framealpha=0.2,
                       labelcolor="white", facecolor="#1a1d2e", edgecolor="#444",
                       fontsize=9)

            ax1.set_title("Eco Estresse com Dobutamina — PAs / PAd / FC por Estágio",
                          color="white", fontsize=11, pad=12)

            fig.tight_layout()
            st.pyplot(fig)

            # Botão de download PNG
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                        facecolor=fig.get_facecolor())
            buf.seek(0)
            st.download_button(
                "🖼️ Baixar gráfico (.png)",
                data=buf,
                file_name=f"eco_estresse_grafico_{nome_pac_ee}.png",
                mime="image/png",
            )
            plt.close(fig)

    st.divider()

    # ── Exportação ───────────────────────────────────────────────────
    tem_dados = any(v for v in st.session_state.valores.values())
    st.subheader("💾 Exportar")
    pac = st.session_state.paciente
    nome_pac = pac.get("Nome","paciente").replace(" ","_").replace("/","-")
    sexo = st.session_state.sexo

    col_csv, col_xls, col_limpar = st.columns([1,1,1])
    with col_csv:
        if tem_dados:
            st.download_button("📄 Baixar CSV",
                data=exportar_csv_bytes(pac, valores_exibir, sexo, estruturado_atual, st.session_state.wmsi_scores),
                file_name=f"eco_{nome_pac}.csv", mime="text/csv",
                use_container_width=True)
        else:
            st.button("📄 Baixar CSV", disabled=True, use_container_width=True)

    with col_xls:
        if tem_dados:
            st.download_button("📊 Baixar Excel",
                data=exportar_excel_bytes(pac, valores_exibir, sexo, estruturado_atual, st.session_state.wmsi_scores),
                file_name=f"eco_{nome_pac}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        else:
            st.button("📊 Baixar Excel", disabled=True, use_container_width=True)

    with col_limpar:
        if st.button("🗑 Limpar Formulário", use_container_width=True):
            st.session_state.valores  = {}
            st.session_state.paciente = {}
            st.rerun()


if __name__ == "__main__":
    main()